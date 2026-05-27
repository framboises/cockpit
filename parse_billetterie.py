#!/usr/bin/env python3
"""
Parse les PDFs comparatifs billetterie et genere un xlsx par annee
avec les sous-totaux, prefixes par le topic_code.

Usage:
  python3 parse_billetterie.py 24M                                  # tous les PDFs dans uploads/
  python3 parse_billetterie.py 24M uploads/mon_fichier.pdf           # un seul PDF
  python3 parse_billetterie.py 24M fichier1.pdf fichier2.pdf         # plusieurs PDFs
"""

import pdfplumber
import re
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill


# ── Utilitaires ──────────────────────────────────────────────────────────

def collect_text(words, x_min, x_max):
    return " ".join(w['text'] for w in words
                    if w['x0'] >= x_min and w['x0'] < x_max).strip()


def parse_int(s):
    s = s.replace("\u20ac", "").replace(",", ".").replace("\xa0", "")
    s = s.replace("\uf0ec", "").strip()
    s = re.sub(r'\s+', '', s)
    if not s or s == '-':
        return None
    try:
        return int(float(s))
    except (ValueError, OverflowError):
        return None


def clean_name(name):
    """Nettoie un nom de produit aire/tribune: retire prix, capacites."""
    name = name.replace("\u20ac", "").strip()
    # Fix espacement lettre a lettre: "2 r o u e s" -> "2 roues"
    def fix_spaced(m):
        return re.sub(r'\s+', '', m.group(0))
    name = re.sub(r'(?<!\w)(\w\s){3,}\w(?!\w)', fix_spaced, name)
    # Retirer prix/tarifs en fin
    name = re.sub(r'\s*\d+\s*,\s*\d+\s*\d*\s*$', '', name)
    # Retirer capacite en debut (3+ chiffres)
    name = re.sub(r'^\d{3}[\d\s]*(?=[A-Z])', '', name)
    name = re.sub(r'^\d{2,}\s+(?=\d\s*[a-z])', '', name)
    # Si juste un nombre, vide
    if re.match(r'^[\d\s,.\u20ac]+$', name.strip()):
        return ''
    return name.strip()


def normalize_name(topic_code, raw_name):
    """Normalise: TOPICCODE_NOMPRODUIT (majuscules, sans accent, sans espace)."""
    import unicodedata
    name = raw_name.strip()
    # Retirer "Carte " prefix
    name = re.sub(r'^Carte\s+', '', name)
    # Majuscules
    name = name.upper()
    # Retirer accents
    name = unicodedata.normalize('NFD', name)
    name = ''.join(c for c in name if unicodedata.category(c) != 'Mn')
    # Remplacer espaces et caracteres speciaux par _
    name = re.sub(r'[^A-Z0-9]+', '_', name)
    # Nettoyer underscores multiples et en debut/fin
    name = re.sub(r'_+', '_', name).strip('_')
    return f"{topic_code} {name}"


# ── Parseur ──────────────────────────────────────────────────────────────

def find_column_boundaries(page):
    """
    Trouve les positions X des colonnes Vendu
    depuis la ligne d'en-tete ("Vendu" et "Commande").
    Retourne (v24_range, v25_range).
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=2)

    vendu_positions = []
    cmd_positions = []

    for w in words:
        if w['top'] > 60:
            continue
        if w['text'] == 'Vendu':
            vendu_positions.append(w['x0'])
        elif w['text'] == 'Commandé':
            cmd_positions.append(w['x0'])

    vendu_positions.sort()
    cmd_positions.sort()

    if len(vendu_positions) >= 2 and len(cmd_positions) >= 2:
        v24 = (vendu_positions[0] - 2, cmd_positions[0])
        v25 = (vendu_positions[1] - 2, cmd_positions[1])
        return v24, v25

    return (381, 418), (670, 706)


def parse_subtotals(page):
    """Extrait tous les sous-totaux + totaux tribunes/aires/supplements/parking."""
    words = page.extract_words(x_tolerance=2, y_tolerance=2)

    v24_range, v25_range = find_column_boundaries(page)

    rows = {}
    for w in words:
        y = round(w['top'] / 3) * 3
        if y not in rows:
            rows[y] = []
        rows[y].append(w)
    for y in rows:
        rows[y].sort(key=lambda w: w['x0'])

    x_split = page.width / 2
    label_24_end = v24_range[0]
    label_25_end = v25_range[0]

    results_24 = []
    results_25 = []

    # Trouver Y du "Total 202X" (fin de la zone ENTREES)
    y_total_entrees = None

    # ── 1. Sous-totaux de la zone ENTREES ──
    for y in sorted(rows.keys()):
        ws = rows[y]
        line = " ".join(w['text'] for w in ws)

        if 'Total 202' in line:
            y_total_entrees = y
            continue

        if 'ous-total' not in line:
            continue

        # Cote 2024
        left = [w for w in ws if w['x0'] < x_split]
        label_24 = collect_text(left, 0, label_24_end)
        m = re.search(r'ous-total\s+(.*)', label_24)
        if m:
            name = m.group(1).strip()
            vendu = parse_int(collect_text(left, *v24_range))
            if name and vendu is not None:
                results_24.append((name, vendu))

        # Cote 2025
        right = [w for w in ws if w['x0'] >= x_split]
        label_25 = collect_text(right, x_split, label_25_end)
        m = re.search(r'ous-total\s+(.*)', label_25)
        if m:
            name = m.group(1).strip()
            vendu = parse_int(collect_text(right, *v25_range))
            if name and vendu is not None:
                results_25.append((name, vendu))

    if not y_total_entrees:
        y_total_entrees = 575

    # ── 2. Supplements (nom sur une ligne, donnees 3px au-dessus) ──
    supp_names = []
    for y in sorted(rows.keys()):
        if y < y_total_entrees or y > y_total_entrees + 35:
            continue
        ws = rows[y]
        margin = collect_text(ws, 200, 280)
        margin = re.sub(r'^SUPP\s*', '', margin).strip()
        if margin and not re.match(r'^[\d\s.,\u20ac%]+$', margin):
            supp_names.append((y, margin))

    for name_y, name in supp_names:
        for offset in [-3, -6, -9]:
            data_y = name_y + offset
            if data_y in rows:
                ws = rows[data_y]
                v24 = parse_int(collect_text(ws, *v24_range))
                v25 = parse_int(collect_text(ws, *v25_range))
                # Fallback: vendu peut etre dans la zone total pour les supplements
                if v24 is None:
                    v24 = parse_int(collect_text(ws, v24_range[1], v24_range[1] + 30))
                if v24 is not None or v25 is not None:
                    if v24 is not None:
                        results_24.append((name, v24))
                    if v25 is not None:
                        results_25.append((name, v25))
                    break

    # ── 3. Tribunes (produits individuels) ──
    # Trouver les bornes Y: entre le header "Capacite" apres ENTREES et le Total tribunes
    y_trib_start = y_total_entrees + 30
    y_trib_end = None
    y_aires_end = None
    # Trouver les lignes "Total" a x<270 (tribunes puis aires)
    total_lines = []
    for y in sorted(rows.keys()):
        if y <= y_total_entrees:
            continue
        ws = rows[y]
        margin_words = [w for w in ws if w['x0'] < 270]
        margin_text = " ".join(w['text'] for w in margin_words).strip()
        if margin_text == 'Total' or (margin_text.startswith('Total') and len(margin_text) < 10):
            total_lines.append(y)
    y_trib_end = total_lines[0] if len(total_lines) >= 1 else y_trib_start + 60
    y_aires_end = total_lines[1] if len(total_lines) >= 2 else y_trib_end + 100

    # Detecter les noms de tribunes dans la marge
    tribune_markers = {}
    for y in sorted(rows.keys()):
        if y < y_trib_start or y > y_trib_end:
            continue
        margin = collect_text(rows[y], 200, 295)
        if 'T16' in margin:
            tribune_markers[y] = 'Tribune T16'
        elif 'T19' in margin:
            tribune_markers[y] = 'Tribune T19'
        elif 'T34' in margin:
            tribune_markers[y] = 'Tribune T34'

    def get_tribune(y):
        name = 'Tribune'
        for ty in sorted(tribune_markers.keys()):
            if ty <= y:
                name = tribune_markers[ty]
        return name

    # Extraire les lignes produit tribunes (celles avec un tarif en euros)
    for y in sorted(rows.keys()):
        if y < y_trib_start or y >= y_trib_end:
            continue
        ws = rows[y]
        line = " ".join(w['text'] for w in ws)
        if 'Capacit' in line or 'Occupation' in line:
            continue
        # Chercher un tarif (xx €) dans la zone tarif 2024 (x ~355-385) ou 2025
        tarif_zone = collect_text(ws, 355, 390)
        if not re.search(r'\d+\s*\u20ac', tarif_zone):
            continue
        v24 = parse_int(collect_text(ws, *v24_range))
        v25 = parse_int(collect_text(ws, *v25_range))
        trib = get_tribune(y)
        if v24 is not None:
            results_24.append((trib, v24))
        if v25 is not None:
            results_25.append((trib, v25))

    # ── 4. Aires d'accueil (produits individuels) ──
    y_aires_start = y_trib_end + 3

    # Detecter les noms d'aires dans la marge
    aire_markers = {}
    for y in sorted(rows.keys()):
        if y < y_aires_start or y > y_aires_end:
            continue
        margin = collect_text(rows[y], 200, 295)
        if 'Houx' in margin or 'Concentration' in margin:
            aire_markers[y] = 'AA Houx'
        elif 'Ouest' in margin:
            aire_markers[y] = 'AA Ouest'
        elif 'Maison' in margin:
            aire_markers[y] = 'AA Maison Blanche'
        elif 'Panorama' in margin:
            aire_markers[y] = 'AA Panorama'
        elif 'Tertre' in margin:
            aire_markers[y] = 'AA Tertre Rouge'

    def get_aire(y):
        name = 'AA'
        for ay in sorted(aire_markers.keys()):
            if ay <= y:
                name = aire_markers[ay]
        return name

    # Extraire les produits aires (Seule, Pack, 2 roues, 4 roues, etc.)
    for y in sorted(rows.keys()):
        if y < y_aires_start or y >= y_aires_end:
            continue
        ws = rows[y]
        line = " ".join(w['text'] for w in ws)
        if 'Total' in line or 'Capacit' in line or 'Occupation' in line or 'Remplissag' in line:
            continue
        if 'en 202' in line:
            continue

        # Nom produit cote 2025 (plus fiable, noms a jour)
        nom_25 = collect_text(ws, 570, 665)
        nom_24 = collect_text(ws, 280, 370)
        # Nettoyer: retirer chiffres purs (capacites)
        if re.match(r'^[\d\s]+$', nom_24): nom_24 = ''
        if re.match(r'^[\d\s]+$', nom_25): nom_25 = ''
        nom_24 = clean_name(nom_24)
        nom_25 = clean_name(nom_25)
        if not nom_24 and not nom_25:
            continue

        v24 = parse_int(collect_text(ws, *v24_range))
        v25 = parse_int(collect_text(ws, *v25_range))

        aire = get_aire(y)
        if v24 is not None and nom_24:
            results_24.append((f"{aire} {nom_24}", v24))
        if v25 is not None and nom_25:
            results_25.append((f"{aire} {nom_25}", v25))

    # ── 5. Parking ──
    for y in sorted(rows.keys()):
        if y < y_aires_end:
            continue
        ws = rows[y]
        margin = collect_text(ws, 200, 270)
        if 'Parking' in margin or ('8' in margin and 'AB' in margin.replace(' ', '')):
            v24 = parse_int(collect_text(ws, *v24_range))
            v25 = parse_int(collect_text(ws, *v25_range))
            if v24 is not None:
                results_24.append(('Parking 8AB', v24))
            if v25 is not None:
                results_25.append(('Parking 8AB', v25))
            break

    return results_24, results_25


def detect_years(page):
    words = page.extract_words(x_tolerance=2, y_tolerance=2)
    text = " ".join(w['text'] for w in sorted(words, key=lambda w: (w['top'], w['x0'])))
    years = re.findall(r'(?:Tarifs|Total|CA\s+\w+)\s+(20\d{2})', text)
    years = sorted(set(years))
    return (int(years[0]), int(years[1])) if len(years) >= 2 else (None, None)


def detect_date_from_page1(page):
    """Detecte la date DD/MM/YYYY en haut a droite de la page 1."""
    words = page.extract_words(x_tolerance=2, y_tolerance=2)
    for w in words:
        if w['top'] < 50:
            m = re.match(r'(\d{2})/(\d{2})/(\d{4})', w['text'])
            if m:
                return m.group(1), m.group(2), int(m.group(3))
    return None, None, None


# ── Variante Reporting Hebdo (5+ pages, tableau pages 4-5) ───────────────

def detect_variant(pdf):
    """Renvoie 'reporting_hebdo', 'lmc' ou 'billetterie' selon la structure du PDF."""
    if len(pdf.pages) >= 5:
        t1 = pdf.pages[0].extract_text() or ''
        if 'Reporting Hebdo' in t1 or 'évolution billetterie' in t1.lower():
            return 'reporting_hebdo'
    if len(pdf.pages) == 3:
        t1 = pdf.pages[0].extract_text() or ''
        if 'LMC' in t1 or 'Etat Hebdo' in t1 or re.search(r'LMC_Comparatif', t1):
            return 'lmc'
    return 'billetterie'


def detect_date_lmc(page):
    """Pour LMC : date DD/MM/YY (année 2 chiffres), peut être n'importe où sur la page."""
    words = page.extract_words(x_tolerance=2, y_tolerance=2)
    # Priorité aux dates les plus en haut
    dates = []
    for w in words:
        m = re.match(r'(\d{2})/(\d{2})/(\d{2})$', w['text'])
        if m:
            dates.append((w['top'], m.group(1), m.group(2), 2000 + int(m.group(3))))
    if dates:
        # Prendre la date la plus en haut (snapshot le plus récent)
        dates.sort(key=lambda d: (d[0], -d[3]))
        _, dd, mm, yyyy = dates[0]
        return dd, mm, yyyy
    return None, None, None


def _group_rows(page, y_tol=3):
    """Regroupe les mots par ligne (Y arrondi) et trie par X."""
    words = page.extract_words(x_tolerance=2, y_tolerance=2)
    rows = {}
    for w in words:
        y = round(w['top'] / y_tol) * y_tol
        rows.setdefault(y, []).append(w)
    for y in rows:
        rows[y].sort(key=lambda w: w['x0'])
    return rows


def detect_years_reporting(pdf):
    """Détecte (annee_N-1, annee_N) depuis les en-têtes des pages 4/5."""
    years = set()
    for pidx in (3, 4):
        if pidx >= len(pdf.pages):
            continue
        text = pdf.pages[pidx].extract_text() or ''
        for m in re.finditer(r'(?:Tarif|Total|CA\s+\w+)\s+(20\d{2})', text):
            years.add(int(m.group(1)))
    if len(years) >= 2:
        ys = sorted(years)
        return ys[0], ys[1]
    return None, None


# Zones X pour Reporting Hebdo (page 4 + haut page 5)
RH_SUBTOT_OLD = (245, 275)
RH_SUBTOT_NEW = (543, 570)
RH_VAR_ZONE = (572, 605)
RH_LABEL_ZONE = (40, 105)


RH_LABEL_ZONE_WIDE = (35, 145)
RH_TARIF_OLD = (125, 188)
RH_TARIF_NEW = (412, 475)


def _is_subtotal_context(ws):
    """Les DEUX côtés TARIF doivent être vides (pas de €, pas de tarif numérique)."""
    t_old = collect_text(ws, *RH_TARIF_OLD)
    t_new = collect_text(ws, *RH_TARIF_NEW)
    if '€' in t_old or '€' in t_new:
        return False
    if re.search(r'\d', t_old) or re.search(r'\d', t_new):
        return False
    return True


# Catégories de sous-totaux Reporting Hebdo connues — utilisées comme liste blanche
# pour accepter les lignes dont un côté TARIF est dirty (cas Essais Jeudi
# où la ligne valeur est fusionnée avec un produit Tarif Invalide)
KNOWN_SUBTOT_LABELS = {
    'Journée Test', 'Essais Mercredi', 'Essais Jeudi', 'Essais 2 Jours',
    'Week-End', 'Dimanche', 'Mulsanne-Arnage', 'Tarif Partenaires', 'Semaine',
}


def _is_known_subtotal_label(label):
    if not label:
        return False
    return label.strip() in KNOWN_SUBTOT_LABELS


def _extract_subtotaux_categories(rows, y_min=100, y_max=1150):
    """Sous-totaux catégorie ventes (Journée Test, Essais X, Week-End, Tarif Partenaires...).
    Stratégie :
      - Ligne avec variation% dans RH_VAR_ZONE
      - Au moins un côté TARIF clean (filtre les produits)
      - Valeur Σ Ventes au moins d'un côté
      - Label trouvé sur ligne idx ou idx+1 ou idx-1 (en marge x ∈ [35,145])"""
    out = []
    keys = sorted(rows.keys())
    seen = set()

    def _clean_label(s):
        s = re.sub(r'^[\d\s.,€%-]+', '', s).strip()
        s = re.sub(r'[\d\s.,€%-]+$', '', s).strip()
        return s

    def _find_label(idx):
        # Zone élargie pour vérifier que la ligne label ne contient pas de tarif (€ ou XX,XX)
        LABEL_PURITY_ZONE = (35, 200)
        for j in (idx, idx+1, idx-1):
            if j < 0 or j >= len(keys):
                continue
            ws_j = rows[keys[j]]
            cand_check = collect_text(ws_j, *LABEL_PURITY_ZONE)
            if '€' in cand_check or re.search(r'\d+\s*[,.]\s*\d{2}', cand_check):
                continue
            cand = _clean_label(collect_text(ws_j, *RH_LABEL_ZONE_WIDE))
            if cand and len(cand) >= 3 and re.search(r'[A-Za-zÀ-ÿ]{3}', cand) and '€' not in cand:
                if cand.lower() in {'tx', 'en 2024', 'en 2025'}:
                    continue
                return cand
        return None

    for i, y in enumerate(keys):
        if y < y_min or y > y_max:
            continue
        ws = rows[y]
        var_text = collect_text(ws, *RH_VAR_ZONE)
        if not re.search(r'-?\d+[,.]?\d*\s*%', var_text):
            continue
        n_old_t = collect_text(ws, *RH_SUBTOT_OLD)
        n_new_t = collect_text(ws, *RH_SUBTOT_NEW)
        n_old = parse_int(n_old_t) if '€' not in n_old_t else None
        n_new = parse_int(n_new_t) if '€' not in n_new_t else None
        if n_old is None and n_new is None:
            continue
        strict_ok = _is_subtotal_context(ws)
        label = _find_label(i)
        if not label:
            continue
        # Acceptation : soit règle stricte (TARIF des 2 côtés vides),
        # soit label dans la liste blanche des catégories connues (cas Essais Jeudi fusionné)
        if not strict_ok and not _is_known_subtotal_label(label):
            continue
        if label in seen:
            continue
        seen.add(label)
        out.append((label, n_old, n_new))
    return out


def _find_y_total_entrees(rows5):
    """Y de la ligne 'Total 2024 ... Total 2025' (TOTAL Entrées)."""
    for y in sorted(rows5):
        if y > 260:
            break
        ws = rows5[y]
        line = " ".join(w['text'] for w in ws)
        if re.search(r'\bTotal\s+20\d{2}\b', line):
            return y
    return None


def _extract_total_entrees(rows5, y):
    """Extrait TOTAL Entrées (vendu+cmd) pour les deux années.
    Cas spécial : sur certains PDFs, le label 'Total 2024 Total 2025' est seul sur Y
    et les valeurs sont sur Y+3."""
    if y is None or y not in rows5:
        return None, None
    keys = sorted(rows5.keys())
    candidates = [rows5[y]]
    try:
        idx = keys.index(y)
        if idx + 1 < len(keys):
            candidates.append(rows5[keys[idx+1]])
    except ValueError:
        pass
    for ws in candidates:
        total_old = parse_int(collect_text(ws, 228, 250))
        total_new = parse_int(collect_text(ws, 525, 547))
        if total_old or total_new:
            return total_old, total_new
    return None, None


def _extract_tribunes(rows5, y_start, y_end):
    """Tribunes individuelles : extraire TOTAL vendu+cmd.
    Cas spécial : si le nom est long ou découpé (T12, T17), les valeurs peuvent être sur la ligne précédente."""
    out = []
    TOTAL_OLD = (228, 250)
    TOTAL_NEW = (525, 547)
    keys = sorted([y for y in rows5 if y_start < y < y_end])
    seen_names = set()
    for i, y in enumerate(keys):
        ws = rows5[y]
        name_full = collect_text(ws, 30, 110)
        # Regex qui capture T_X avec suffixes (bis, ter) ou T29bis collé
        m = re.match(r'^(T\d+(?:\s*bis|\s*ter)?(?:bis|ter)?)\b', name_full)
        if not m:
            continue
        full_id = re.sub(r'\s+', '', m.group(1)).lower()  # ex: "T3bis"
        if full_id in seen_names:
            continue
        total_old = parse_int(collect_text(ws, *TOTAL_OLD))
        total_new = parse_int(collect_text(ws, *TOTAL_NEW))
        # Fallback : ligne précédente
        if total_old is None and total_new is None and i > 0:
            ws_prev = rows5[keys[i-1]]
            total_old = parse_int(collect_text(ws_prev, *TOTAL_OLD))
            total_new = parse_int(collect_text(ws_prev, *TOTAL_NEW))
        if total_old is None and total_new is None:
            continue
        clean = re.sub(r'\s*-\s*', ' ', name_full).strip()
        out.append((clean, total_old, total_new))
        seen_names.add(full_id)
    return out


# ── Variante LMC (Etat Hebdo 3 pages, page 3 = comparatif 2023 vs 2025) ─

# Zones X pour LMC (page portrait 595 wide)
LMC_NAME_X = (30, 135)
LMC_TOTAL_OLD = (211, 215.5)
LMC_TOTAL_NEW = (363, 370)


def _find_y_total_lmc(rows, y_max=200):
    """Y de la ligne 'Total 2023 ... Total 2025' (Total Entrées)."""
    for y in sorted(rows):
        if y > y_max:
            break
        line = " ".join(w['text'] for w in rows[y])
        if re.search(r'Total\s+20\d{2}', line):
            return y
    return None


def _sum_values_in_zone(words, x_min, x_max):
    """Somme les valeurs entières dans la zone X, regroupées par micro-ligne (top arrondi à 0.5).
    Chaque micro-ligne contient un nombre (potentiellement avec séparateur de milliers : '54 499').
    Plusieurs micro-lignes superposées (Plein Tarif + Tarif ACO empilés) sont sommées."""
    by_row = {}
    for w in words:
        if not (x_min <= w['x0'] < x_max):
            continue
        key = round(w['top'] * 2) / 2
        by_row.setdefault(key, []).append(w)
    total = 0
    found = False
    for ws in by_row.values():
        ws.sort(key=lambda w: w['x0'])
        joined = ' '.join(w['text'] for w in ws)
        v = parse_int(joined)
        if v is not None:
            total += v
            found = True
    return total if found else None


def parse_lmc(pdf):
    """Parse LMC Etat Hebdo : extrait Total Entrées, Tribunes, Aires + leurs totaux.

    Stratégie : extraction fine (y_tolerance=0.5), détection dynamique des colonnes TOTAL,
    SOMMATION des valeurs entre tribunes/aires (chaque tribune a plusieurs sous-tarifs empilés)."""
    page3 = pdf.pages[2]
    words = page3.extract_words(x_tolerance=1, y_tolerance=0.5)

    # Détection dynamique : positions X des en-têtes "TOTAL" et années depuis 'Tarifs 20XX'
    # (top < 60 pour couvrir les variants où header est à top ~ 48)
    header_words = [w for w in words if w['top'] < 60]
    total_x_positions = sorted(w['x0'] for w in header_words if w['text'] == 'TOTAL')
    years_in_header = sorted(set(int(m.group(1)) for w in header_words
                                  for m in [re.match(r'^(20\d{2})$', w['text'])] if m))

    if len(total_x_positions) >= 2 and len(years_in_header) >= 2:
        # Deux colonnes TOTAL — calibrer zones X autour de chacune
        x_total_old_center = total_x_positions[0]
        x_total_new_center = total_x_positions[1]
        # Le label "TOTAL" est au-dessus de la colonne ; la zone des valeurs commence à peu près au même x
        # Width typique 8-12 (assez pour "62 246" mais pas la colonne adjacente)
        X_OLD = (x_total_old_center - 4, x_total_old_center + 8)
        X_NEW = (x_total_new_center - 4, x_total_new_center + 8)
        year_old, year_new = years_in_header[0], years_in_header[1]
    else:
        # Fallback : positions fixes
        X_OLD = LMC_TOTAL_OLD
        X_NEW = LMC_TOTAL_NEW
        year_old = years_in_header[0] if years_in_header else None
        year_new = years_in_header[1] if len(years_in_header) >= 2 else None

    out_old, out_new = [], []

    def words_in_y(top_min, top_max):
        return [w for w in words if top_min <= w['top'] <= top_max]

    def row_words(top, half=1.0):
        return [w for w in words if abs(w['top'] - top) <= half]

    # Pré-calcul : tous les "Total 20XX" rows (peut être à différents Y selon le PDF)
    total_year_rows = []  # liste de (top, t_o, t_n)
    seen_tops = set()
    for w in words:
        if w['text'] == 'Total' and w['x0'] < 200 and w['top'] > 60:
            same_top = round(w['top'])
            if same_top in seen_tops:
                continue
            # Année doit être EXACTEMENT sur la même micro-row (top diff < 0.4)
            same = [w2['text'] for w2 in words if abs(w2['top'] - w['top']) < 0.4]
            if any(re.match(r'^20\d{2}$', t) for t in same):
                seen_tops.add(same_top)
                # Mais pour les valeurs, on accepte la ligne entière (les valeurs peuvent être sur Y+0.5)
                rwords = row_words(w['top'])
                t_o = _sum_values_in_zone(rwords, *X_OLD)
                t_n = _sum_values_in_zone(rwords, *X_NEW)
                total_year_rows.append((w['top'], t_o, t_n))
    total_year_rows.sort()

    # 1) Total Entrées = premier "Total 20XX" trouvé
    y_total_e = None
    if total_year_rows:
        y_total_e, t_o, t_n = total_year_rows[0]
        if t_o: out_old.append(('Total Entrees', t_o))
        if t_n: out_new.append(('Total Entrees', t_n))

    # 2) Tribunes : marqueurs Tx (Y > Y_total_e pour éviter faux positifs)
    # x marge variable selon PDF : étendre à [70, 130]
    tribunes = []
    for w in words:
        if 70 < w['x0'] < 130 and w['top'] > (y_total_e or 60) + 10 and re.match(r'^T\d+(?:bis|ter)?$', w['text']):
            tribunes.append((w['top'], w['text']))
    tribunes.sort()

    y_last_trib = tribunes[-1][0] if tribunes else 0

    # 3) Total Tribunes : premier "Total" en marge après la dernière tribune (sans 20XX)
    y_total_trib = None
    if tribunes:
        for w in sorted(words, key=lambda w: w['top']):
            if w['text'] == 'Total' and w['x0'] < 150 and w['top'] > y_last_trib + 3:
                same = [w2['text'] for w2 in words if abs(w2['top'] - w['top']) < 0.4]
                if any(re.match(r'^20\d{2}$', t) for t in same):
                    continue
                y_total_trib = w['top']
                rwords = row_words(w['top'])
                t_o = _sum_values_in_zone(rwords, *X_OLD)
                t_n = _sum_values_in_zone(rwords, *X_NEW)
                if t_o: out_old.append(('Total Tribunes', t_o))
                if t_n: out_new.append(('Total Tribunes', t_n))
                break

    # 4) Tribunes individuelles : sommer entre tribune i et tribune i+1 (Y boundaries = milieux)
    def _trib_name(top, tx_label):
        # Nom : mots strictement sur la même micro-row que le marqueur Tx, x ∈ [100, 130]
        # + le nom (mot alpha) qui suit immédiatement (Siko, Wimille, etc.)
        same = [w for w in words if abs(w['top'] - top) < 0.6 and 100 < w['x0'] < 140
                and not re.match(r'^\d+$', w['text']) and w['text'] != '-']
        same.sort(key=lambda w: w['x0'])
        full = ' '.join(w['text'] for w in same)
        return full.strip() or tx_label

    for i, (top, tname) in enumerate(tribunes):
        full = _trib_name(top, tname)
        # Y range : milieu entre tribunes pour éviter overlap
        prev_top = tribunes[i-1][0] if i > 0 else top - 12
        next_top = tribunes[i+1][0] if i+1 < len(tribunes) else (y_total_trib if y_total_trib else top + 12)
        y_start = (prev_top + top) / 2 + 0.5 if i > 0 else top - 4
        y_end = (top + next_top) / 2 - 0.5
        rwords = words_in_y(y_start, y_end)
        t_o = _sum_values_in_zone(rwords, *X_OLD)
        t_n = _sum_values_in_zone(rwords, *X_NEW)
        if t_o is not None: out_old.append((full, t_o))
        if t_n is not None: out_new.append((full, t_n))

    # 5) Aires : détecter les noms (mots alpha en marge entre y_total_trib et y_total_aires)
    if y_total_trib:
        # Trouver d'abord y_total_aires (le 'Total' avec la plus grande somme après y_total_trib)
        totals_after_trib = []
        for w in words:
            if w['text'] == 'Total' and 100 < w['x0'] < 150 and w['top'] > y_total_trib + 30:
                rwords = row_words(w['top'])
                t_o = _sum_values_in_zone(rwords, *X_OLD)
                t_n = _sum_values_in_zone(rwords, *X_NEW)
                totals_after_trib.append((w['top'], t_o, t_n))

        y_total_aires = None
        if totals_after_trib:
            big = [t for t in totals_after_trib if (t[1] or 0) + (t[2] or 0) > 1000]
            if big:
                chosen = max(big, key=lambda x: (x[1] or 0) + (x[2] or 0))
                y_total_aires = chosen[0]
                if chosen[1]: out_old.append(('Total Aires', chosen[1]))
                if chosen[2]: out_new.append(('Total Aires', chosen[2]))

        # Aires individuelles
        if y_total_aires:
            # Détecter le X de la marge des noms d'aires (varie selon PDF: 102 pour S-1, ~54 pour S-13)
            # → on prend le X le plus fréquent parmi les mots alpha en marge entre y_total_trib et y_total_aires
            from collections import Counter
            margin_x_counts = Counter()
            for w in words:
                if not (y_total_trib + 3 < w['top'] < y_total_aires - 2):
                    continue
                if not (30 <= w['x0'] < 130):
                    continue
                if re.match(r'^[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\-]{2,}$', w['text']):
                    margin_x_counts[round(w['x0'])] += 1
            if not margin_x_counts:
                anchor_x = 102
            else:
                anchor_x = margin_x_counts.most_common(1)[0][0]

            aire_anchors = []
            for w in words:
                if not (y_total_trib + 3 < w['top'] < y_total_aires - 2):
                    continue
                if not (anchor_x - 2 <= w['x0'] < anchor_x + 3):
                    continue
                if re.match(r'^[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\-]{2,}$', w['text']):
                    blacklist = {'total', 'tarif', 'plein', 'tarifs', 'vendu', 'commandé',
                                 'capacité', 'occupation', 'tribune', 'capacit',
                                 'gratuit', 'inclus', 'salarié', 'etudiant', 'pmr',
                                 'contrôleur', 'individuel', 'jeudi', 'vendredi',
                                 'samedi', 'dimanche', 'mercredi', 'mardi', 'cible',
                                 'groupe', 'remplissag', 'paddock', 'pitlounge'}
                    if w['text'].lower() in blacklist:
                        continue
                    aire_anchors.append((w['top'], w['x0'], w['text']))

            aire_anchors.sort()
            # Pour chaque ancre, construire le nom complet en prenant les mots adjacents
            # de la même micro-row (top ±0.5) à partir de l'ancre, en excluant chiffres et symboles
            grouped = []
            for top_a, x_a, txt_a in aire_anchors:
                name_words = [w for w in words if abs(w['top'] - top_a) < 0.6
                              and anchor_x - 2 <= w['x0'] < anchor_x + 60
                              and re.match(r'^[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\-]+$', w['text'])]
                name_words.sort(key=lambda w: w['x0'])
                full_name = ' '.join(w['text'] for w in name_words)
                grouped.append((top_a, full_name))

            seen = set()
            unique_aires = []
            for top, name in grouped:
                if name.lower() in seen or len(name) < 3:
                    continue
                seen.add(name.lower())
                unique_aires.append((top, name))

            # Pour chaque aire, sommer dans la zone Y jusqu'à l'aire suivante
            for k, (top, name) in enumerate(unique_aires):
                y_start = top - 2
                y_end = unique_aires[k+1][0] - 1 if k+1 < len(unique_aires) else y_total_aires - 1
                rwords = words_in_y(y_start, y_end)
                t_o = _sum_values_in_zone(rwords, *X_OLD)
                t_n = _sum_values_in_zone(rwords, *X_NEW)
                if t_o is None and t_n is None:
                    continue
                if t_o is not None: out_old.append((name, t_o))
                if t_n is not None: out_new.append((name, t_n))

    # 6) Total intermédiaire (Green Tickets / Suppléments) : Total entre y_total_e et la première tribune
    if y_total_e and tribunes:
        intermediate_totals = []
        for w in words:
            if w['text'] == 'Total' and 100 < w['x0'] < 150 and y_total_e + 30 < w['top'] < tribunes[0][0] - 5:
                rwords = row_words(w['top'])
                t_o = _sum_values_in_zone(rwords, *X_OLD)
                t_n = _sum_values_in_zone(rwords, *X_NEW)
                if (t_o or 0) + (t_n or 0) > 100:
                    intermediate_totals.append((w['top'], t_o, t_n))
        if intermediate_totals:
            _, t_o, t_n = intermediate_totals[-1]  # le plus proche des tribunes
            if t_o: out_old.append(('Total Section Intermediaire', t_o))
            if t_n: out_new.append(('Total Section Intermediaire', t_n))

    return year_old, year_new, out_old, out_new


# ── Extraction produits hospitalité (Cartes 24, Accréditations) ─────────

CARTES_NAMES = [
    'Pavillon 24 (ex-Arena 24)',
    'Carte Pitlounge 24',
    'Le Mans Spirit Club',
    'Gold Expérience',
    'Peninsula 24',
    'Pavillon 24',
    'Square 24',
    'Terrace 24',
    'Starter 24',
    'Panoramic',
    'Premium',
]


def _detect_carte_name(text):
    """Retourne le nom de Carte si le texte commence par un nom connu."""
    s = text.strip()
    for c in CARTES_NAMES:
        if s.startswith(c):
            return c
    return None


def _clean_product_name(text, carte=None):
    """Nettoie un nom de produit : retire la Carte préfixe puis ne garde que les mots alpha (et tirets)."""
    if not text:
        return ''
    s = text.strip()
    if carte and s.startswith(carte):
        s = s[len(carte):].strip()
    # Garder uniquement les fragments commençant par une lettre (continue avec lettres/tirets/slash/point)
    parts = re.findall(r"[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\-./]*", s)
    # Filtrer fragments d'un seul caractère parasites
    parts = [p for p in parts if len(p) >= 2 or p.lower() in ('a',)]
    return ' '.join(parts).strip()


def _extract_cartes_products(rows4, rows5, y_total_e):
    """Extrait les produits hospitalité (Cartes 24) en pages 4-5 jusqu'à Accréditations.
    Retourne [(nom_old_complet, t_old, nom_new_complet, t_new), ...]"""
    out = []

    NOM_ZONE_OLD = (30, 162)
    NOM_ZONE_OLD_WIDE = (30, 190)
    TARIF_OLD = (158, 190)
    TOTAL_OLD = (228, 250)
    NOM_ZONE_NEW = (335, 450)
    NOM_ZONE_NEW_WIDE = (335, 478)
    TARIF_NEW = (450, 478)
    TOTAL_NEW = (525, 547)

    def _proc(rows, y_start, y_end, current_carte):
        keys = sorted([y for y in rows if y_start <= y <= y_end])
        for i, y in enumerate(keys):
            ws = rows[y]
            nom_left = collect_text(ws, *NOM_ZONE_OLD)
            detected = _detect_carte_name(nom_left)
            if detected:
                current_carte = detected
            if not current_carte:
                continue
            tarif_o = collect_text(ws, *TARIF_OLD)
            tarif_n = collect_text(ws, *TARIF_NEW)
            t_old = parse_int(collect_text(ws, *TOTAL_OLD)) if '€' in tarif_o else None
            t_new = parse_int(collect_text(ws, *TOTAL_NEW)) if '€' in tarif_n else None
            if t_old is None and t_new is None:
                continue
            nom_o = _clean_product_name(collect_text(ws, *NOM_ZONE_OLD_WIDE), current_carte)
            nom_n = _clean_product_name(collect_text(ws, *NOM_ZONE_NEW_WIDE), current_carte)
            # Fallback : ligne précédente (cas Gold Expérience où nom sur Y-3)
            if not nom_o and not nom_n and i > 0:
                ws_prev = rows[keys[i-1]]
                nom_o = _clean_product_name(collect_text(ws_prev, *NOM_ZONE_OLD_WIDE), current_carte)
                nom_n = _clean_product_name(collect_text(ws_prev, *NOM_ZONE_NEW_WIDE), current_carte)
            if not nom_o and not nom_n:
                continue
            full_o = f"{current_carte} {nom_o}".strip() if nom_o else current_carte
            full_n = f"{current_carte} {nom_n}".strip() if nom_n else current_carte
            out.append((full_o, t_old, full_n, t_new))
        return current_carte

    # Trouver Y de "Accréditations" sur page 5
    y_accredi = None
    for y in sorted(rows5):
        if y > y_total_e:
            break
        if 'Accréditations' in " ".join(w['text'] for w in rows5[y]):
            y_accredi = y
            break

    current = _proc(rows4, 830, 1170, None)
    if y_accredi:
        _proc(rows5, 30, y_accredi - 5, current)
    return out


def _extract_accreditations(rows5, y_accredi_start, y_total_e):
    """Extrait les produits Accréditations (Enceinte Générale, Pitlounge, Paddock, etc.) entre 'Accréditations' et 'Total 202X'."""
    out = []
    NOM_ZONE_OLD = (30, 190)
    TARIF_OLD = (158, 190)
    TOTAL_OLD = (228, 250)
    NOM_ZONE_NEW = (335, 478)
    TARIF_NEW = (450, 478)
    TOTAL_NEW = (525, 547)
    keys = sorted([y for y in rows5 if y_accredi_start < y < y_total_e])
    for i, y in enumerate(keys):
        ws = rows5[y]
        tarif_o = collect_text(ws, *TARIF_OLD)
        tarif_n = collect_text(ws, *TARIF_NEW)
        t_old = parse_int(collect_text(ws, *TOTAL_OLD)) if '€' in tarif_o else None
        t_new = parse_int(collect_text(ws, *TOTAL_NEW)) if '€' in tarif_n else None
        if t_old is None and t_new is None:
            continue
        nom_o = _clean_product_name(collect_text(ws, *NOM_ZONE_OLD))
        nom_n = _clean_product_name(collect_text(ws, *NOM_ZONE_NEW))
        # Fallback ligne précédente si nom vide
        if not nom_o and not nom_n and i > 0:
            ws_prev = rows5[keys[i-1]]
            nom_o = _clean_product_name(collect_text(ws_prev, *NOM_ZONE_OLD))
            nom_n = _clean_product_name(collect_text(ws_prev, *NOM_ZONE_NEW))
        if not nom_o and not nom_n:
            continue
        full_o = f"Accreditation {nom_o}".strip() if nom_o else 'Accreditation'
        full_n = f"Accreditation {nom_n}".strip() if nom_n else 'Accreditation'
        out.append((full_o, t_old, full_n, t_new))
    return out


def _extract_total_section(rows5, y_label):
    """Lit le total à Y_label (ou Y_label - 3 en fallback) pour les colonnes TOTAL_OLD/NEW."""
    if y_label is None:
        return None, None
    TOTAL_OLD = (228, 250)
    TOTAL_NEW = (525, 547)
    candidates = []
    keys = sorted(rows5.keys())
    if y_label in rows5:
        candidates.append(rows5[y_label])
    # Ligne immédiatement précédente (Y_label - 3)
    try:
        idx = keys.index(y_label)
        if idx > 0:
            candidates.append(rows5[keys[idx-1]])
    except ValueError:
        pass
    for ws in candidates:
        to = parse_int(collect_text(ws, *TOTAL_OLD))
        tn = parse_int(collect_text(ws, *TOTAL_NEW))
        if to or tn:
            return to, tn
    return None, None


def _extract_aires(rows5, y_start, y_end):
    """Aires : extraire pour chaque aire un total vendu+cmd 2024 et 2025.
    3 cas :
     a) ligne résumé sans € (ex: Arnage Y=465) — totaux x ∈ [108,165] et [395,460], occupation%
     b) ligne unique avec € (ex: Houx Y=615, aire à 1 produit) — totaux x ∈ [228,250] et [525,547]
     c) nom isolé sans valeurs (ex: Clos Fleuri Y=492) — totaux pris sur la ligne produit qui suit"""
    out = []
    NAME_X = (35, 110)
    OCC_OLD_X = (260, 295)
    OCC_NEW_X = (560, 605)
    SUMMARY_OLD = (108, 165)
    SUMMARY_NEW = (395, 460)
    PRODUCT_OLD = (228, 250)
    PRODUCT_NEW = (525, 547)

    keys = sorted(rows5.keys())
    seen = set()
    for idx, y in enumerate(keys):
        if y <= y_start or y >= y_end:
            continue
        ws = rows5[y]
        name = collect_text(ws, *NAME_X)
        if not name or len(name) > 35 or len(name) < 3:
            continue
        if name.lower().startswith(('total', 'tx', 'tarif', 'capacit', 'occupation',
                                     'remplissag', 'en 20', 'plein ', 'green ',
                                     'accreditation', 'accréditation', 'bracelet',
                                     'coursives', 'le mans', 'paddock', 'pitlounge',
                                     'enceinte', 'club ', 'salarié', 'salarie',
                                     'france billet')):
            continue
        # Nom doit contenir au moins 3 lettres consécutives
        if not re.search(r'[A-Za-zÀ-ÿ]{3}', name):
            continue
        line = " ".join(w['text'] for w in ws)
        occ = collect_text(ws, *OCC_OLD_X) + " " + collect_text(ws, *OCC_NEW_X)
        has_occ = bool(re.search(r'\d+[,.]?\d*\s*%', occ))

        total_old = None
        total_new = None
        if has_occ:
            # Cas a) ou b)
            if '€' in line:
                total_old = parse_int(collect_text(ws, *PRODUCT_OLD))
                total_new = parse_int(collect_text(ws, *PRODUCT_NEW))
            else:
                total_old = parse_int(collect_text(ws, *SUMMARY_OLD))
                total_new = parse_int(collect_text(ws, *SUMMARY_NEW))
        else:
            # Cas c) — nom isolé : utiliser la ligne produit qui suit (jusqu'à 4 lignes plus loin)
            for j in range(idx+1, min(idx+5, len(keys))):
                ws_next = rows5[keys[j]]
                line_next = " ".join(w['text'] for w in ws_next)
                if '€' not in line_next:
                    continue
                to = parse_int(collect_text(ws_next, *PRODUCT_OLD))
                tn = parse_int(collect_text(ws_next, *PRODUCT_NEW))
                if to is not None or tn is not None:
                    total_old, total_new = to, tn
                    break

        if total_old is None and total_new is None:
            continue
        if name in seen:
            continue
        seen.add(name)
        out.append((name, total_old, total_new))
    return out


def parse_reporting_hebdo(pdf):
    """Parse PDF Reporting Hebdo. Renvoie (year_old, year_new, results_old, results_new).
    results = [(nom_produit_prefixe, vendu_net), ...]"""
    year_old, year_new = detect_years_reporting(pdf)
    if not year_old:
        return None, None, [], []

    page4 = pdf.pages[3]
    page5 = pdf.pages[4]
    rows4 = _group_rows(page4)
    rows5 = _group_rows(page5)

    out_old, out_new = [], []

    # 1) Sous-totaux catégories ventes (page 4)
    for nom, n_old, n_new in _extract_subtotaux_categories(rows4, y_min=100, y_max=1150):
        if n_old is not None: out_old.append((nom, n_old))
        if n_new is not None: out_new.append((nom, n_new))

    # 2) Total Entrées (page 5)
    y_total_e = _find_y_total_entrees(rows5)
    # 2bis) Sous-totaux catégorie en haut page 5 (avant Total Entrées)
    if y_total_e:
        for nom, n_old, n_new in _extract_subtotaux_categories(rows5, y_min=30, y_max=y_total_e - 3):
            if n_old is not None: out_old.append((nom, n_old))
            if n_new is not None: out_new.append((nom, n_new))

    # 2ter) Produits Cartes 24 (pages 4 + 5 jusqu'à Accréditations)
    if y_total_e:
        for n_o, t_o, n_n, t_n in _extract_cartes_products(rows4, rows5, y_total_e):
            if t_o is not None and n_o: out_old.append((n_o, t_o))
            if t_n is not None and n_n: out_new.append((n_n, t_n))

    # 2quater) Produits Accréditations (page 5 entre 'Accréditations' et 'Total 202X')
    if y_total_e:
        y_accredi = None
        for y in sorted(rows5):
            if y > y_total_e:
                break
            if 'Accréditations' in " ".join(w['text'] for w in rows5[y]):
                y_accredi = y
                break
        if y_accredi:
            for n_o, t_o, n_n, t_n in _extract_accreditations(rows5, y_accredi, y_total_e):
                if t_o is not None and n_o: out_old.append((n_o, t_o))
                if t_n is not None and n_n: out_new.append((n_n, t_n))

    t_old, t_new = _extract_total_entrees(rows5, y_total_e)
    if t_old is not None: out_old.append(('Total Entrees', t_old))
    if t_new is not None: out_new.append(('Total Entrees', t_new))

    # 3) Trouver Y des "Total" tribunes et aires
    y_total_trib = None
    y_total_aires = None
    if y_total_e:
        # Total tribunes : ligne avec 'Total' dans la zone label (x~80-120), Y entre y_total_e+50 et fin
        for y in sorted(rows5):
            if y <= y_total_e + 50:
                continue
            ws = rows5[y]
            margin_text = collect_text(ws, 75, 130)
            if margin_text.strip().startswith('Total'):
                if y_total_trib is None:
                    y_total_trib = y
                elif y_total_aires is None:
                    y_total_aires = y
                    break

    # 4) Tribunes individuelles
    if y_total_e and y_total_trib:
        for nom, t_old, t_new in _extract_tribunes(rows5, y_total_e + 50, y_total_trib):
            if t_old is not None: out_old.append((nom, t_old))
            if t_new is not None: out_new.append((nom, t_new))
        # Total Tribunes (valeurs souvent sur ligne Y-3 par rapport au label)
        t_old, t_new = _extract_total_section(rows5, y_total_trib)
        if t_old is not None: out_old.append(('Total Tribunes', t_old))
        if t_new is not None: out_new.append(('Total Tribunes', t_new))

    # 5) Aires individuelles
    if y_total_trib and y_total_aires:
        for nom, a_old, a_new in _extract_aires(rows5, y_total_trib + 5, y_total_aires):
            if a_old is not None: out_old.append((nom, a_old))
            if a_new is not None: out_new.append((nom, a_new))
        # Total Aires
        t_old, t_new = _extract_total_section(rows5, y_total_aires)
        if t_old is not None: out_old.append(('Total Aires', t_old))
        if t_new is not None: out_new.append(('Total Aires', t_new))

    return year_old, year_new, out_old, out_new


# ── XLSX ─────────────────────────────────────────────────────────────────

def write_xlsx(products, year, dd, mm, output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "productReceipts"

    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.cell(row=1, column=2, value="Produit").font = hdr_font
    ws.cell(row=1, column=2).fill = hdr_fill
    ws.cell(row=1, column=2).border = border
    ws.cell(row=1, column=3, value="Nb net de ventes (Total)").font = hdr_font
    ws.cell(row=1, column=3).fill = hdr_fill
    ws.cell(row=1, column=3).border = border

    total = 0
    for i, (nom, vendu) in enumerate(products, 2):
        ws.cell(row=i, column=2, value=nom).border = border
        c = ws.cell(row=i, column=3, value=vendu)
        c.border = border
        c.number_format = '#,##0'
        total += (vendu or 0)

    row = len(products) + 2
    ws.cell(row=row, column=2, value="Grand total").font = Font(bold=True)
    ws.cell(row=row, column=2).border = border
    c = ws.cell(row=row, column=3, value=total)
    c.font = Font(bold=True)
    c.border = border
    c.number_format = '#,##0'

    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25

    filename = f"Query_{year}{mm}{dd}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ── Traitement ───────────────────────────────────────────────────────────

def process_pdf(pdf_path, topic_code, output_dir):
    print(f"\nTraitement: {os.path.basename(pdf_path)}")

    pdf = pdfplumber.open(pdf_path)
    if len(pdf.pages) < 3:
        print(f"  SKIP: {len(pdf.pages)} pages")
        pdf.close()
        return []

    variant = detect_variant(pdf)

    if variant == 'lmc':
        dd, mm, date_year = detect_date_lmc(pdf.pages[0])
    else:
        dd, mm, date_year = detect_date_from_page1(pdf.pages[0])

    if not dd:
        print("  WARN: date non trouvee sur page 1")
        pdf.close()
        return []

    if variant == 'reporting_hebdo':
        year_old, year_new, st_old, st_new = parse_reporting_hebdo(pdf)
        if not year_old:
            year_old, year_new = date_year - 1, date_year
        print(f"  Variant: Reporting Hebdo")
    elif variant == 'lmc':
        year_old, year_new, st_old, st_new = parse_lmc(pdf)
        if not year_old:
            year_old, year_new = date_year - 2, date_year  # LMC bisannuel
        print(f"  Variant: LMC Etat Hebdo")
    else:
        year_old, year_new = detect_years(pdf.pages[2])
        if not year_old:
            year_old, year_new = date_year - 1, date_year
        st_old, st_new = parse_subtotals(pdf.pages[2])

    # Prefixer les noms avec le topic_code
    st_new = [(normalize_name(topic_code, n), v) for n, v in st_new]
    st_old = [(normalize_name(topic_code, n), v) for n, v in st_old]

    print(f"  Date: {dd}/{mm}/{date_year}")
    print(f"  {year_new}: {len(st_new)} sous-totaux | {year_old}: {len(st_old)} sous-totaux")
    for name, v in st_new:
        print(f"    {name:<45} {v:>10}")

    files = []
    if st_new:
        f = write_xlsx(st_new, year_new, dd, mm, output_dir)
        print(f"  -> {os.path.basename(f)}")
        files.append(f)
    if st_old:
        f = write_xlsx(st_old, year_old, dd, mm, output_dir)
        print(f"  -> {os.path.basename(f)}")
        files.append(f)

    pdf.close()
    return files


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('-')]

    if not args:
        print("Usage: python3 parse_billetterie.py TOPIC_CODE [fichier.pdf ...]")
        print("  Ex:  python3 parse_billetterie.py 24M")
        print("       python3 parse_billetterie.py 24M uploads/fichier.pdf")
        sys.exit(1)

    topic_code = args[0]
    pdf_args = args[1:]

    script_dir = os.path.dirname(os.path.abspath(__file__))
    uploads_dir = os.path.join(script_dir, "uploads")
    output_dir = os.path.join(uploads_dir, "xlsx_output")
    os.makedirs(output_dir, exist_ok=True)

    if pdf_args:
        pdfs = []
        for p in pdf_args:
            path = p if os.path.isabs(p) else os.path.join(os.getcwd(), p)
            if os.path.exists(path):
                pdfs.append(path)
            else:
                print(f"Fichier introuvable: {p}")
    else:
        pdfs = sorted([
            os.path.join(uploads_dir, f)
            for f in os.listdir(uploads_dir)
            if f.endswith('.pdf') and 'Billetterie' in f
        ])

    if not pdfs:
        print("Aucun PDF billetterie trouve")
        return

    print(f"Topic code: {topic_code}")
    print(f"{len(pdfs)} PDF(s) a traiter")

    all_files = []
    for pdf_path in pdfs:
        try:
            files = process_pdf(pdf_path, topic_code, output_dir)
            all_files.extend(files)
        except Exception as e:
            print(f"  ERREUR: {e}")
            import traceback
            traceback.print_exc()

    print(f"\nTermine! {len(all_files)} fichiers xlsx dans {output_dir}")


if __name__ == "__main__":
    main()
