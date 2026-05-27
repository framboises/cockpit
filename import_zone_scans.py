"""
Import des scans entree/sortie par zone depuis un xlsx 24h du Mans.

Structure source (uploads/zone-complet 24hdu mans.xlsx) :
- Ligne 1 : datetime (uniquement sur les colonnes Entree, fusion visuelle avec Sortie)
- Ligne 2 : 'Entree' / 'Sortie'
- Ligne 3 : libelle 'SPACE_CODE - Identifiant' (ignore)
- Lignes 4+ : nom de zone en colonne A, valeurs scans par pas de 15 min
- Derniere colonne : 'Total' (ignoree)
- Lignes 'ENCEINTE GENERALE' et 'Total' ignorees

Cible : collection `parking_scans` de la base titan_dev.
Un document par zone, intervalles embarques :
  {
    event: '24h_du_mans', year: 2025, zone: 'AA ARNAGE',
    period_start, period_end,
    total_entree, total_sortie,
    intervals: [{ts, entree, sortie}, ...]
  }
"""

import os
import sys
from datetime import datetime
import openpyxl
from pymongo import MongoClient

XLSX_PATH = '/Users/framboises/Dropbox/ACO/TITAN/cockpit/uploads/zone-complet 24hdu mans.xlsx'
EVENT = '24h_du_mans'
YEAR = 2025
COLLECTION = 'parking_scans'
SKIP_ZONES = {'ENCEINTE GENERALE', 'Total'}

MONGO_URI = os.getenv('MONGO_URI', 'mongodb://localhost:27017/')
DB_NAME = 'titan_dev'


def build_column_map(header_row, sense_row):
    """Retourne dict { col_index (1-based): (ts, 'entree'|'sortie') }."""
    col_map = {}
    last_ts = None
    for c in range(2, len(header_row) + 1):
        h1 = header_row[c - 1]
        sense = sense_row[c - 1] if c - 1 < len(sense_row) else None
        if h1 == 'Total':
            continue
        if isinstance(h1, datetime):
            last_ts = h1
        if sense in ('Entree', 'Entrée'):
            ts = h1 if isinstance(h1, datetime) else last_ts
            if ts is None:
                continue
            col_map[c] = (ts, 'entree')
        elif sense == 'Sortie':
            if last_ts is None:
                continue
            col_map[c] = (last_ts, 'sortie')
    return col_map


def build_documents_from_rows(rows_iter, col_map):
    """Itere les lignes (1-based row index, tuple values) et construit les docs."""
    docs = []
    for r_idx, row in rows_iter:
        if r_idx < 4:
            continue
        name = row[0]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if name in SKIP_ZONES:
            continue
        intervals_by_ts = {}
        total_e = 0
        total_s = 0
        for c, (ts, kind) in col_map.items():
            if c - 1 >= len(row):
                continue
            v = row[c - 1]
            if v is None:
                continue
            try:
                n = int(v)
            except (TypeError, ValueError):
                try:
                    n = int(float(v))
                except (TypeError, ValueError):
                    continue
            slot = intervals_by_ts.setdefault(ts, {'ts': ts, 'entree': 0, 'sortie': 0})
            slot[kind] = n
            if kind == 'entree':
                total_e += n
            else:
                total_s += n
        intervals = sorted(intervals_by_ts.values(), key=lambda x: x['ts'])
        if not intervals:
            continue
        docs.append({
            'event': EVENT,
            'year': YEAR,
            'zone': name,
            'period_start': intervals[0]['ts'],
            'period_end': intervals[-1]['ts'],
            'total_entree': total_e,
            'total_sortie': total_s,
            'intervals_count': len(intervals),
            'intervals': intervals,
            'source_file': os.path.basename(XLSX_PATH),
            'imported_at': datetime.utcnow(),
        })
    return docs


def main():
    if not os.path.isfile(XLSX_PATH):
        print(f'ERREUR : fichier introuvable {XLSX_PATH}', file=sys.stderr)
        sys.exit(1)
    print(f'Chargement {XLSX_PATH} ...')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    sense_row = None
    data_rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            header_row = row
        elif r_idx == 2:
            sense_row = row
        elif r_idx >= 4:
            data_rows.append((r_idx, row))

    col_map = build_column_map(header_row, sense_row)
    print(f'  Colonnes de donnees identifiees : {len(col_map)} '
          f'(entree={sum(1 for _,k in col_map.values() if k=="entree")}, '
          f'sortie={sum(1 for _,k in col_map.values() if k=="sortie")})')

    docs = build_documents_from_rows(data_rows, col_map)
    print(f'  Documents construits : {len(docs)}')

    wb.close()

    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    col = db[COLLECTION]

    col.create_index([('event', 1), ('year', 1), ('zone', 1)], unique=True)
    col.create_index([('event', 1), ('year', 1)])

    inserted = 0
    updated = 0
    for d in docs:
        res = col.replace_one(
            {'event': d['event'], 'year': d['year'], 'zone': d['zone']},
            d,
            upsert=True,
        )
        if res.upserted_id is not None:
            inserted += 1
        elif res.modified_count:
            updated += 1

    total = col.count_documents({'event': EVENT, 'year': YEAR})
    print(f'OK. Inserts={inserted}, updates={updated}. Total docs en base ({EVENT}/{YEAR}) : {total}')

    sample = col.find_one(
        {'event': EVENT, 'year': YEAR},
        {'zone': 1, 'total_entree': 1, 'total_sortie': 1, 'intervals_count': 1,
         'period_start': 1, 'period_end': 1, '_id': 0},
    )
    print(f'  Exemple : {sample}')


if __name__ == '__main__':
    main()
