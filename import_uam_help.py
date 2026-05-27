"""
Identifie l'aide UAM (Unite d'Appui Mobile) et les renforts PDA sur les portes.

Logique :
- Les PDAs apparaissant dans la ligne (zone=ENCEINTE GENERALE, porte=UAM) du
  xlsx POC sont des "UAM PDAs". Ils se deplacent sur les portes en renfort.
- Quand un UAM PDA apparait dans une autre porte (= row porte != UAM), on
  enregistre ses scans comme "aide UAM" pour cette porte.
- Pour les portes a tripodes (place_config.tripodes=true dans la collection
  portes), les PDAs non-UAM utilises sont aussi enregistres comme "renfort PDA"
  (signal d'un debordement des tripodes).

Stockage : dans porte_scans, ajout des champs :
- uam_help : per_hour + per_day (nb pda + nb scans)
- pda_renfort : (portes tripodes seulement) per_hour + per_day
"""

import os
from collections import defaultdict
from datetime import datetime
import openpyxl
from pymongo import MongoClient

XLSX_PATH = os.getenv(
    'PORTE_XLSX',
    '/Users/framboises/Dropbox/ACO/TITAN/cockpit/uploads/zone-complet 24hdu mans-poc.xlsx',
)
EVENT = '24h_du_mans'
YEAR = 2025

MONGO_URI = os.getenv('MONGO_URI', 'mongodb://localhost:27017/')
DB_NAME = 'titan_dev'


def build_column_map(header_row, sense_row):
    col_map = {}
    last_ts = None
    for c in range(3, len(header_row)):
        h1 = header_row[c]
        sense = sense_row[c] if c < len(sense_row) else None
        if h1 == 'Total': continue
        if isinstance(h1, datetime): last_ts = h1
        if sense in ('Entree', 'Entrée'):
            ts = h1 if isinstance(h1, datetime) else last_ts
            if ts: col_map[c] = (ts, 'entree')
        elif sense == 'Sortie':
            if last_ts: col_map[c] = (last_ts, 'sortie')
    return col_map


def main():
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]

    # 1. Charge le flag tripodes depuis collection portes
    tripode_portes = set()
    geo = db['portes'].find_one() or {}
    for f in geo.get('features', []) or []:
        p = f.get('properties') or {}
        nm = (p.get('Name') or '').strip().upper()
        pc = p.get('place_config') or {}
        if nm and pc.get('tripodes'):
            tripode_portes.add(nm)
    print(f'Portes tripodes (geojson) : {sorted(tripode_portes)}')

    # 2. Charge le xlsx
    print(f'Chargement {XLSX_PATH}')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_row, sense_row = rows[0], rows[1]
    col_map = build_column_map(header_row, sense_row)

    # 3. Identifie les UAM PDAs
    uam_pdas = set()
    last_zone = last_porte = None
    for r in rows[3:]:
        if r[0]: last_zone = r[0]
        if r[1]: last_porte = r[1]
        name = r[2]
        if not name: continue
        zu = (last_zone or '').upper()
        pu = (last_porte or '').upper()
        if zu == 'UAM' or pu == 'UAM':
            if name.startswith('PDA'):
                uam_pdas.add(name)
    print(f'UAM PDAs identifies : {len(uam_pdas)}')

    # 4. Pour chaque (porte, device, ts), on cumule les scans
    # uam_data[porte][hour] = {pda_set, scan_count}
    # pda_renfort[porte][hour] = idem (portes tripodes uniquement, hors UAM)
    uam_data = defaultdict(lambda: defaultdict(lambda: {'pdas': set(), 'scans': 0}))
    pda_renfort_data = defaultdict(lambda: defaultdict(lambda: {'pdas': set(), 'scans': 0}))

    last_zone = last_porte = None
    for r in rows[3:]:
        if r[0]: last_zone = r[0]
        if r[1]: last_porte = r[1]
        name = r[2]
        if not name: continue
        zu = (last_zone or '').upper()
        pu = (last_porte or '').upper()
        if zu != 'ENCEINTE GENERALE':
            continue
        if pu == 'UAM':
            continue  # ligne UAM elle-meme, pas une porte cible
        if not name.startswith('PDA'):
            continue  # tripodes ne sont pas comptes en renfort PDA
        is_uam = name in uam_pdas
        is_tripode_porte = pu in tripode_portes
        if not is_uam and not is_tripode_porte:
            continue  # rien a enregistrer pour ce cas
        # parcours des scans
        for c, (ts, kind) in col_map.items():
            if c >= len(r): continue
            v = r[c]
            if v is None: continue
            try: n = int(v)
            except (TypeError, ValueError): continue
            if n <= 0: continue
            hour_key = ts.replace(minute=0).strftime('%Y-%m-%dT%H:00')
            if is_uam:
                slot = uam_data[last_porte][hour_key]
                slot['pdas'].add(name)
                slot['scans'] += n
            if is_tripode_porte and not is_uam:
                slot = pda_renfort_data[last_porte][hour_key]
                slot['pdas'].add(name)
                slot['scans'] += n

    wb.close()
    print(f'Portes avec aide UAM detectee : {len(uam_data)}')
    print(f'Portes tripodes avec renfort PDA detecte : {len(pda_renfort_data)}')

    # 5. Construit les structures par jour aussi (resume)
    def to_dict(per_hour):
        out_h = {}
        per_day = defaultdict(lambda: {'pdas': set(), 'scans': 0, 'hours': 0})
        for hour_key, slot in per_hour.items():
            out_h[hour_key] = {
                'pda_count': len(slot['pdas']),
                'scan_count': slot['scans'],
            }
            day = hour_key[:10]
            d = per_day[day]
            d['pdas'].update(slot['pdas'])
            d['scans'] += slot['scans']
            d['hours'] += 1
        out_d = {
            day: {
                'pda_count': len(v['pdas']),
                'scan_count': v['scans'],
                'active_hours': v['hours'],
            }
            for day, v in per_day.items()
        }
        return out_h, out_d

    # 6. Upsert dans porte_scans
    portes_in_scan = set(d['porte'] for d in db['porte_scans'].find(
        {'event': EVENT, 'year': YEAR}, {'porte': 1}))

    updates = 0
    for porte in portes_in_scan:
        u_per_hour, u_per_day = to_dict(uam_data.get(porte, {}))
        p_per_hour, p_per_day = to_dict(pda_renfort_data.get(porte, {}))
        update = {
            'uam_help': {
                'per_hour': u_per_hour,
                'per_day': u_per_day,
                'total_scans': sum(v['scan_count'] for v in u_per_hour.values()),
            },
        }
        if porte.upper() in tripode_portes:
            update['pda_renfort'] = {
                'per_hour': p_per_hour,
                'per_day': p_per_day,
                'total_scans': sum(v['scan_count'] for v in p_per_hour.values()),
            }
        res = db['porte_scans'].update_one(
            {'event': EVENT, 'year': YEAR, 'porte': porte},
            {'$set': update},
        )
        if res.matched_count:
            updates += 1

    print(f'Portes mises a jour : {updates}')

    # 7. Echantillon
    sample = db['porte_scans'].find_one({'event': EVENT, 'year': YEAR,
                                          'uam_help.total_scans': {'$gt': 1000}})
    if sample:
        print(f'\nExemple : {sample["porte"]}')
        u = sample['uam_help']
        print(f'  UAM total scans : {u["total_scans"]}, jours touches : {len(u["per_day"])}')
        # top 5 hours
        top5 = sorted(u['per_hour'].items(), key=lambda x: -x[1]['scan_count'])[:5]
        for h, v in top5:
            print(f'  {h} : {v["pda_count"]} PDA UAM, {v["scan_count"]} scans')


if __name__ == '__main__':
    main()
