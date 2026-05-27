"""
Import des scans entree/sortie par PORTE depuis le xlsx POC (hierarchique).

Structure source :
- Col 1 = zone (ex ENCEINTE GENERALE)
- Col 2 = porte (ex PORTE NORD PIETONS)
- Col 3 = device (PDA.xxx ou TRI-xxx)
- Col 4+ = colonnes scans 15 min (pair Entree/Sortie sous datetime fusionne)
- Lignes 1-3 = en-tetes, lignes 4+ = donnees

On filtre la zone ENCEINTE GENERALE (= "portes du circuit") et on agrege par
(zone, porte) en sommant les scans de tous les devices. Resultat ecrit dans
`titan_dev.porte_scans` avec le meme shape que `parking_scans` plus champs
specifiques (porte, device_count, pda_count, tripode_count).
"""

import os
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl
from pymongo import MongoClient

XLSX_PATH = os.getenv(
    'PORTE_XLSX',
    '/Users/framboises/Dropbox/ACO/TITAN/cockpit/uploads/zone-complet 24hdu mans-poc.xlsx',
)
EVENT = os.getenv('SCAN_EVENT', '24h_du_mans')
YEAR = int(os.getenv('SCAN_YEAR', '2025'))
COLLECTION = 'porte_scans'
TARGET_ZONES = {'ENCEINTE GENERALE'}

MONGO_URI = os.getenv('MONGO_URI', 'mongodb://localhost:27017/')
DB_NAME = 'titan_dev'


def build_column_map(header_row, sense_row):
    """Mappe chaque colonne de donnees -> (datetime, 'entree'|'sortie')."""
    col_map = {}
    last_ts = None
    for c in range(3, len(header_row)):  # cols 0,1,2 = zone, porte, device
        h1 = header_row[c]
        sense = sense_row[c] if c < len(sense_row) else None
        if h1 == 'Total':
            continue
        if isinstance(h1, datetime):
            last_ts = h1
        if sense in ('Entree', 'Entrée'):
            ts = h1 if isinstance(h1, datetime) else last_ts
            if ts is None:
                continue
            col_map[c] = (ts, 'entree')
        elif sense == 'Sortie':
            if last_ts is None:
                continue
            col_map[c] = (last_ts, 'sortie')
    return col_map


def main():
    if not os.path.isfile(XLSX_PATH):
        print(f'ERREUR : fichier introuvable {XLSX_PATH}', file=sys.stderr)
        sys.exit(1)
    print(f'Chargement {XLSX_PATH} ...')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_row, sense_row = rows[0], rows[1]
    col_map = build_column_map(header_row, sense_row)
    print(f'  Colonnes data identifiees : {len(col_map)}')

    porte_data = defaultdict(lambda: defaultdict(lambda: {'entree': 0, 'sortie': 0}))
    porte_devices = defaultdict(lambda: {'devices': set(), 'pdas': set(), 'tripodes': set()})
    last_zone = last_porte = None
    for r in rows[3:]:
        if r[0]:
            last_zone = r[0]
        if r[1]:
            last_porte = r[1]
        name = r[2]
        if not name:
            continue
        if last_zone not in TARGET_ZONES:
            continue
        key = (last_zone, last_porte)
        porte_devices[key]['devices'].add(name)
        if name.startswith('PDA'):
            porte_devices[key]['pdas'].add(name)
        elif name.startswith('TRI'):
            porte_devices[key]['tripodes'].add(name)
        for c, (ts, kind) in col_map.items():
            if c >= len(r):
                continue
            v = r[c]
            if v is None:
                continue
            try:
                n = int(v)
            except (TypeError, ValueError):
                try:
                    n = int(float(v))
                except (TypeError, ValueError):
                    continue
            porte_data[key][ts][kind] += n

    wb.close()
    print(f'  Portes uniques (zone={list(TARGET_ZONES)[0]}) : {len(porte_data)}')

    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    col = db[COLLECTION]
    col.create_index([('event', 1), ('year', 1), ('zone', 1), ('porte', 1)], unique=True)
    col.create_index([('event', 1), ('year', 1)])

    inserted = updated = 0
    for (zone, porte), intervals_dict in sorted(porte_data.items()):
        intervals = sorted(
            ({'ts': ts, **vals} for ts, vals in intervals_dict.items()),
            key=lambda x: x['ts'],
        )
        if not intervals:
            continue
        total_e = sum(it['entree'] for it in intervals)
        total_s = sum(it['sortie'] for it in intervals)
        dev = porte_devices[(zone, porte)]
        doc = {
            'event': EVENT,
            'year': YEAR,
            'zone': zone,
            'porte': porte,
            'period_start': intervals[0]['ts'],
            'period_end': intervals[-1]['ts'],
            'total_entree': total_e,
            'total_sortie': total_s,
            'intervals_count': len(intervals),
            'intervals': intervals,
            'device_count': len(dev['devices']),
            'pda_count': len(dev['pdas']),
            'tripode_count': len(dev['tripodes']),
            'source_file': os.path.basename(XLSX_PATH),
            'imported_at': datetime.utcnow(),
        }
        res = col.replace_one(
            {'event': EVENT, 'year': YEAR, 'zone': zone, 'porte': porte},
            doc, upsert=True,
        )
        if res.upserted_id is not None:
            inserted += 1
        elif res.modified_count:
            updated += 1

    total = col.count_documents({'event': EVENT, 'year': YEAR})
    print(f'OK. Inserts={inserted}, updates={updated}. Total docs en base : {total}')


if __name__ == '__main__':
    main()
