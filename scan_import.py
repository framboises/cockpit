"""
Import des scans Skidata (xlsx hierarchique) vers `historique_controle`.

Ce module est le coeur de la chaine « scans » pilotee depuis /scan-report :
il parse le fichier Excel, rattache chaque unite de scan a une feature
geographique (`_id_feature`), et construit les trois documents Mongo
(`complet`, `frequentation`, `portes`).

Fonctions pures : aucune dependance a Flask, aucun acces reseau. La base est
toujours passee en argument (jamais `titan_dev` en dur, contrairement aux
anciens scripts import_*.py qui ne tournaient que sur le poste de dev).

Format xlsx attendu (cf. uploads/zone-complet-24HM-2024.xlsx) :
- Ligne 1 : datetime, uniquement sur la 1re colonne de chaque groupe fusionne
- Ligne 2 : sens `Entree` | `Sortie` | `Autre`
- Ligne 3 : 'SPACE_CODE - Identifiant' (ignoree)
- Colonnes 0/1/2 : zone | porte | device, avec report vers le bas
- Lignes 4+ : entiers, pas de 15 min, creneaux vides omis (non zero-remplis)
- La colonne 'Total' et la ligne zone == 'Total' sont ignorees

Le sens `Autre` correspond a un PDA non configure en entree/sortie (pas a une
categorie de scan). Il est compte dans `portes.scan_count` mais stocke a part
dans `complet` et exclu du calcul de presents : sans direction, pas de solde.
"""

import os
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

# La zone qui porte les acces du perimetre du circuit. Ses « portes » sont les
# portes d'enceinte ; toutes les autres zones sont des zones (tribunes,
# parkings, aires d'accueil, hospitalites...).
ENCEINTE = 'ENCEINTE GENERALE'

# Libelles a ignorer en colonne zone (ligne de totaux en fin de tableau).
SKIP_ZONES = {'TOTAL'}

SENS_ENTREE = {'ENTREE', 'ENTRÉE'}
SENS_SORTIE = {'SORTIE'}
SENS_AUTRE = {'AUTRE'}

# Collections geojson interrogeables pour le rattachement d'une unite de scan.
FEATURE_COLLECTIONS = ('portes', 'terrains', 'hospitalites', 'tribunes')

# Champ portant le libelle lisible, variable selon la collection.
FEATURE_NAME_FIELD = {
    'portes': 'Name',
    'terrains': 'Name',
    'hospitalites': 'NOM',
    'tribunes': 'NOM',
}

# Collection des corrections manuelles nom de scan -> feature.
OVERRIDES_COLLECTION = 'scan_feature_overrides'

# Prefixes de bruit a retirer avant rapprochement d'un nom de zone : le xlsx
# prefixe ses zones (`AA ARNAGE`, `P OUEST`) la ou le geojson ne le fait pas.
_ZONE_PREFIX_RE = re.compile(r'^(?:AA|P|PK|PKG|PARKING|CAMPING|AIRE|AIRE D ACCUEIL)\s+')


class ScanImportError(Exception):
    """Erreur d'import exposee au front. `code` est l'identifiant machine."""

    def __init__(self, code, status=400, detail=None):
        super().__init__(code)
        self.code = code
        self.status = status
        self.detail = detail


# ---------------------------------------------------------------------------
# Normalisation des libelles
# ---------------------------------------------------------------------------

def normalize_label(value):
    """Casse, accents et ponctuation replies pour comparer deux libelles.

    'MAISON DES HUNAUDIÈRES' et 'Maison des Hunaudieres' donnent la meme
    chaine ; 'PORTE MAISON-BLANCHE' et 'PORTE MAISON BLANCHE' aussi.
    """
    if not value:
        return ''
    txt = unicodedata.normalize('NFKD', str(value))
    txt = txt.encode('ascii', 'ignore').decode('ascii')
    txt = re.sub(r'[^A-Za-z0-9]+', ' ', txt.upper())
    return txt.strip()


# Variantes orthographiques observees entre le xlsx, les documents curates et
# le geojson, pour un meme acces physique. Appliquees des DEUX cotes de la
# comparaison, donc symetriques et sans perte : 'PORTAIL HOUX 5' rejoint
# 'PORTE HOUX 5' (curate -> 67813fd18789a5f5421af36c), 'PORTE CONCIERGERIE'
# rejoint 'PORTAIL CONCIERGERIE' (geojson).
_PORTE_ALIASES = (
    (re.compile(r'\bPORTAIL\b'), 'PORTE'),
    (re.compile(r'\bVEHICULES\b'), 'VEHICULE'),
    (re.compile(r'\bPIETONS\b'), 'PIETON'),
)


def canonical_porte(name):
    """Libelle de porte replie sur ses variantes orthographiques connues."""
    txt = normalize_label(name)
    for pattern, repl in _PORTE_ALIASES:
        txt = pattern.sub(repl, txt)
    return re.sub(r'\s+', ' ', txt).strip()


def _zone_variants(name):
    """Variantes de rapprochement pour un nom de zone (avec/sans prefixe)."""
    base = normalize_label(name)
    out = [base]
    stripped = _ZONE_PREFIX_RE.sub('', base).strip()
    if stripped and stripped != base:
        out.append(stripped)
    return out


def _tribune_number(name):
    """'TRIBUNE 03 BIS/TER' -> '03 BIS/TER' ; None si ce n'est pas une tribune."""
    m = re.match(r'^TRIBUNE\s+(.+)$', normalize_label(name))
    return m.group(1).strip() if m else None


def _tribune_keys(raw):
    """Cles de rapprochement d'un numero de tribune.

    `ANCIEN 2025` melange les formats ('01', '3', '03 bis'), d'ou la
    normalisation du zero-padding en plus de la forme brute.
    """
    if raw is None:
        return set()
    norm = normalize_label(raw)
    if not norm:
        return set()
    keys = {norm}
    m = re.match(r'^(\d+)(.*)$', norm)
    if m:
        num, suffix = m.group(1), m.group(2).strip()
        for form in (str(int(num)), '%02d' % int(num)):
            keys.add((form + ' ' + suffix).strip())
    return keys


# ---------------------------------------------------------------------------
# Parsing du classeur
# ---------------------------------------------------------------------------

def _build_column_map(header_row, sense_row, first_data_col):
    """Colonne -> (datetime, 'e'|'s'|'a').

    Le datetime n'est porte que par la premiere colonne de chaque groupe
    fusionne, qui peut etre `Autre`, `Entree` ou `Sortie` selon les fichiers :
    on propage donc le dernier datetime rencontre (meme logique que
    import_porte_scans.py, etendue au sens `Autre`).
    """
    col_map = {}
    last_ts = None
    for idx in range(first_data_col, len(sense_row)):
        head = header_row[idx] if idx < len(header_row) else None
        if isinstance(head, datetime):
            last_ts = head
        sense = sense_row[idx]
        if sense is None:
            continue
        key = normalize_label(sense)
        if key in SENS_ENTREE:
            kind = 'e'
        elif key in SENS_SORTIE:
            kind = 's'
        elif key in SENS_AUTRE:
            kind = 'a'
        else:
            continue  # colonne 'Total' ou en-tete inattendu
        if last_ts is None:
            continue
        col_map[idx] = (last_ts, kind)
    return col_map


def _classify_device(name):
    """'pda' | 'tripode' | 'autre' d'apres le prefixe du device."""
    norm = normalize_label(name)
    if norm.startswith('PDA'):
        return 'pda'
    if norm.startswith('TRI'):
        return 'tripode'
    return 'autre'


def parse_scan_xlsx(path):
    """Parse le classeur et retourne les unites de scan agregees.

    Retourne un dict :
        {
          'units': [ {key, name, kind, zone, porte, devices, device_count,
                      pda_count, tripode_count, total_entree, total_sortie,
                      total_autre, slots: {datetime: {'e','s','a'}}} ],
          'period_start', 'period_end', 'slot_count', 'has_autre',
          'source_file', 'zone_count', 'porte_count',
        }

    Les unites de type 'zone' agregent toutes les portes de la zone ; les
    unites de type 'porte' sont les portes de ENCEINTE GENERALE.
    """
    if not os.path.isfile(path):
        raise ScanImportError('fichier_introuvable', 404)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ScanImportError('xlsx_illisible', 400, str(exc))

    try:
        ws = wb[wb.sheetnames[0]]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(rows) < 4:
        raise ScanImportError('xlsx_trop_court', 400)

    header_row, sense_row = rows[0], rows[1]
    col_map = _build_column_map(header_row, sense_row, first_data_col=3)
    if not col_map:
        raise ScanImportError('xlsx_sans_colonnes_scan', 400)

    # Agregation en deux passes : d'abord par (zone, porte, device) brut, pour
    # pouvoir ensuite regrouper au niveau zone ou au niveau porte.
    raw = defaultdict(lambda: {'devices': [], 'slots': defaultdict(lambda: {'e': 0, 's': 0, 'a': 0})})
    zone = porte = None
    seen_devices = set()

    # Detail par boitier, conserve uniquement pour les portes de l'enceinte :
    # c'est le grain necessaire au calcul de l'aide UAM et des renforts, que
    # l'agregation par unite fait disparaitre. {(porte, device): {heure: scans}}
    device_hours = defaultdict(lambda: defaultdict(int))
    uam_devices = set()

    for row in rows[3:]:
        z = row[0] if len(row) > 0 else None
        p = row[1] if len(row) > 1 else None
        dev = row[2] if len(row) > 2 else None
        if z not in (None, ''):
            zone = z
            porte = None  # nouvelle zone : la porte reprend au premier libelle
        if p not in (None, ''):
            porte = p
        if zone is None or normalize_label(zone) in SKIP_ZONES:
            continue
        if porte is None:
            continue

        bucket = raw[(str(zone).strip(), str(porte).strip())]
        if dev not in (None, ''):
            dev_name = str(dev).strip()
            marker = (str(zone).strip(), str(porte).strip(), dev_name)
            if marker not in seen_devices:
                seen_devices.add(marker)
                bucket['devices'].append(dev_name)

        dev_name = str(dev).strip() if dev not in (None, '') else None
        is_enceinte = normalize_label(zone) == normalize_label(ENCEINTE)
        porte_norm = normalize_label(porte)
        # Un boitier rattache a la ligne UAM est un PDA mobile : on le repere
        # ici pour pouvoir, plus loin, reconnaitre ses passages sur les autres
        # portes comme du renfort et non comme du personnel fixe.
        if dev_name and (porte_norm == 'UAM' or normalize_label(zone) == 'UAM'):
            if _classify_device(dev_name) == 'pda':
                uam_devices.add(dev_name)

        for idx, (ts, kind) in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val in (None, ''):
                continue
            try:
                n = int(val)
            except (TypeError, ValueError):
                continue
            bucket['slots'][ts][kind] += n
            if (is_enceinte and dev_name and n > 0
                    and porte_norm != 'UAM'
                    and _classify_device(dev_name) == 'pda'):
                hour = ts.replace(minute=0, second=0, microsecond=0)
                device_hours[(str(porte).strip(), dev_name)][hour] += n

    if not raw:
        raise ScanImportError('xlsx_sans_donnees', 400)

    units = []
    # Portes : chaque porte de l'enceinte generale est une unite.
    for (z, p), data in raw.items():
        if normalize_label(z) != normalize_label(ENCEINTE):
            continue
        units.append(_make_unit(name=p, kind='porte', zone=z, porte=p,
                                devices=data['devices'], slots=data['slots']))

    # Zones : agregation de toutes les portes de la zone (hors enceinte).
    by_zone = defaultdict(lambda: {'devices': [], 'slots': defaultdict(lambda: {'e': 0, 's': 0, 'a': 0})})
    for (z, p), data in raw.items():
        if normalize_label(z) == normalize_label(ENCEINTE):
            continue
        agg = by_zone[z]
        agg['devices'].extend(data['devices'])
        for ts, counts in data['slots'].items():
            for k in ('e', 's', 'a'):
                agg['slots'][ts][k] += counts[k]
    for z, data in by_zone.items():
        units.append(_make_unit(name=z, kind='zone', zone=z, porte=None,
                                devices=data['devices'], slots=data['slots']))

    units.sort(key=lambda u: (u['kind'], u['name']))

    all_ts = sorted({ts for u in units for ts in u['slots']})
    if not all_ts:
        raise ScanImportError('xlsx_sans_creneaux', 400)

    return {
        'units': units,
        'period_start': all_ts[0],
        'period_end': all_ts[-1],
        'slot_count': len(all_ts),
        'has_autre': any(u['total_autre'] for u in units),
        'source_file': os.path.basename(path),
        'zone_count': sum(1 for u in units if u['kind'] == 'zone'),
        'porte_count': sum(1 for u in units if u['kind'] == 'porte'),
        'device_hours': {k: dict(v) for k, v in device_hours.items()},
        'uam_devices': sorted(uam_devices),
    }


def _make_unit(name, kind, zone, porte, devices, slots):
    """Assemble une unite de scan a partir de ses creneaux bruts."""
    devices = sorted(set(d for d in devices if d))
    classes = [_classify_device(d) for d in devices]
    clean_slots = {ts: dict(c) for ts, c in slots.items()
                   if c['e'] or c['s'] or c['a']}
    return {
        'key': (kind, name),
        'name': str(name).strip(),
        'kind': kind,
        'zone': str(zone).strip() if zone else None,
        'porte': str(porte).strip() if porte else None,
        'devices': devices,
        'device_count': len(devices),
        'pda_count': classes.count('pda'),
        'tripode_count': classes.count('tripode'),
        # FLEX / FX / VG / FG : 10 boitiers du fichier 2024 ne sont ni PDA ni
        # tripode. Sans ce compteur ils disparaissaient de l'inventaire.
        'other_count': classes.count('autre'),
        'total_entree': sum(c['e'] for c in clean_slots.values()),
        'total_sortie': sum(c['s'] for c in clean_slots.values()),
        'total_autre': sum(c['a'] for c in clean_slots.values()),
        'slots': clean_slots,
    }


def guess_year(parsed):
    """Annee deduite des dates internes du fichier (majoritaire)."""
    counts = defaultdict(int)
    for u in parsed['units']:
        for ts in u['slots']:
            counts[ts.year] += 1
    return max(counts.items(), key=lambda kv: kv[1])[0] if counts else None


# ---------------------------------------------------------------------------
# Resolution des features geographiques
# ---------------------------------------------------------------------------

def _load_geo_index(db):
    """Index des features : {collection: {norm_name: [(_id_feature, label)]}}.

    Les tribunes sont en plus indexees sur leur numero historique
    (`ANCIEN 2025` / `NUMERO 2026`), seul pont vers les libelles `TRIBUNE nn`
    du xlsx : les noms de tribunes changent d'une edition a l'autre.
    """
    index = {}
    tribune_by_num = defaultdict(list)
    for coll in FEATURE_COLLECTIONS:
        field = FEATURE_NAME_FIELD[coll]
        by_name = defaultdict(list)
        doc = db[coll].find_one() or {}
        for feat in doc.get('features') or []:
            props = feat.get('properties') or {}
            fid = props.get('_id_feature')
            if not fid:
                continue
            label = props.get(field)
            if label:
                key = canonical_porte(label) if coll == 'portes' else normalize_label(label)
                by_name[key].append((fid, str(label).strip()))
            if coll == 'tribunes':
                for key_field in ('ANCIEN 2025', 'NUMERO 2026'):
                    for key in _tribune_keys(props.get(key_field)):
                        tribune_by_num[key].append((fid, str(label or '').strip()))
        index[coll] = by_name
    index['_tribune_by_num'] = tribune_by_num
    return index


def _harvest_known_doors(db):
    """Nom de porte -> _id_feature, recolte sur les documents `portes` existants.

    Ces documents ont ete curates au fil des editions : ils constituent la
    source de rattachement la plus fiable, et la seule qui connaisse les
    libelles historiques du xlsx (`PORTE NORD` -> PORTE NORD PIETONS).
    """
    harvest = {}
    ambiguous = set()
    for doc in db['historique_controle'].find({'type': 'portes'}, {'doors': 1}):
        for door in doc.get('doors') or []:
            name, fid = door.get('name'), door.get('doors_id')
            if not name or not fid:
                continue
            key = canonical_porte(name)
            if key in harvest and harvest[key] != fid:
                ambiguous.add(key)
            harvest[key] = fid
    for key in ambiguous:
        harvest.pop(key, None)
    return harvest


def _load_overrides(db, event=None, year=None):
    """Corrections manuelles : {(kind, norm_name): {_id_feature, collection}}.

    Un override sans event/year s'applique a tous les imports (cas normal :
    les libelles se repetent d'une edition a l'autre) ; un override cible sur
    un couple (event, year) est prioritaire.
    """
    out = {}
    try:
        cur = db[OVERRIDES_COLLECTION].find({})
    except Exception:
        return out
    for doc in cur:
        key = (doc.get('kind'), normalize_label(doc.get('scan_name')))
        targeted = doc.get('event') is not None and doc.get('year') is not None
        if targeted and (doc.get('event') != event or doc.get('year') != year):
            continue
        if key in out and not targeted:
            continue  # un override cible deja pose est prioritaire
        out[key] = {
            '_id_feature': doc.get('_id_feature'),
            'feature_collection': doc.get('feature_collection'),
            'category': doc.get('category'),
            'ignored': doc.get('ignored'),
            'targeted': targeted,
        }
    return out


# ---------------------------------------------------------------------------
# Categorie d'unite
# ---------------------------------------------------------------------------
#
# La categorie n'est PAS le rattachement geographique. Elle dit ce qu'est
# l'unite pour l'exploitation, et elle pilote trois choses dans le rapport :
# le regroupement de la liste laterale, l'encadre « Vue par categorie », et
# surtout la CAPACITE retenue (650 personnes/h pour un lieu pietonnier, 250
# vehicules/h pour un parking) — donc l'effectif recommande.
#
# Elle etait jusqu'ici devinee par prefixe de nom, avec les conventions
# 24H AUTOS en dur (`TRIBUNE `, `AA `, `P `). Les editions qui ne les suivent
# pas (BEAUSEJOUR, PARKING OUEST, KARTING SUD) n'atterrissaient dans aucune
# categorie et heritaient de la capacite vehicule par defaut.

UNIT_CATEGORIES = [
    {'id': 'porte', 'label': 'Porte d\'enceinte', 'flow': 'personnes'},
    {'id': 'tribune', 'label': 'Tribune', 'flow': 'personnes'},
    {'id': 'aire_accueil', 'label': 'Aire d\'accueil', 'flow': 'vehicules'},
    {'id': 'parking', 'label': 'Parking', 'flow': 'vehicules'},
    {'id': 'paddock', 'label': 'Paddock', 'flow': 'personnes'},
    {'id': 'hospitalite', 'label': 'Hospitalite', 'flow': 'personnes'},
    {'id': 'autre', 'label': 'Autre', 'flow': 'vehicules'},
]
CATEGORY_IDS = {c['id'] for c in UNIT_CATEGORIES}

# Libelles connus qui ne suivent aucune convention de prefixe.
_CATEGORY_BY_NAME = {
    'PADDOCKS': 'paddock',
    'MAISON DES HUNAUDIERES': 'hospitalite',
    'WELCOME': 'hospitalite',
    'VISITES MUSEE': 'hospitalite',
}

# Collection geo -> categorie, quand le nom ne tranche pas. C'est plus fiable
# qu'un prefixe : une unite rattachee a une feature de `tribunes` EST une
# tribune, quel que soit son libelle.
_CATEGORY_BY_COLLECTION = {
    'portes': 'porte',
    'tribunes': 'tribune',
    'hospitalites': 'hospitalite',
}


def guess_category(name, kind, feature_collection=None):
    """Categorie proposee pour une unite. Toujours une valeur, jamais None.

    Ordre : le nom d'abord (le plus explicite pour l'exploitant), puis le
    rattachement geographique, puis un repli par nature d'unite.
    """
    if kind == 'porte':
        return 'porte'
    n = (name or '').strip().upper()
    if n in _CATEGORY_BY_NAME:
        return _CATEGORY_BY_NAME[n]
    if n.startswith('TRIBUNE'):
        return 'tribune'
    if n.startswith('AA ') or 'AIRE D' in n:
        return 'aire_accueil'
    if n.startswith('P ') or n.startswith('PARKING'):
        return 'parking'
    if 'PADDOCK' in n:
        return 'paddock'
    if 'HOSPITALIT' in n:
        return 'hospitalite'
    by_coll = _CATEGORY_BY_COLLECTION.get(feature_collection)
    if by_coll:
        return by_coll
    return 'autre'


def _suggest_for(unit, geo, exclude_ids, limit=6):
    """Rapprochements approximatifs, pour alimenter le mapping manuel.

    Sans cela une unite comme `PORTE NORD` (le geojson ne connait que
    PORTE NORD PIETONS / VL / BIS) n'offrirait aucune piste dans l'UI.
    Score de Jaccard sur les tokens, meme approche que `_auto_match_location`
    (app.py:2633).
    """
    tokens = set(normalize_label(unit['name']).split())
    if not tokens:
        return []
    colls = ('portes',) if unit['kind'] == 'porte' else ('terrains', 'hospitalites', 'tribunes')
    scored = []
    for coll in colls:
        for norm_name, entries in geo[coll].items():
            other = set(norm_name.split())
            if not other:
                continue
            inter = tokens & other
            if not inter:
                continue
            score = len(inter) / float(len(tokens | other))
            if tokens <= other or other <= tokens:
                score += 0.5
            for fid, label in entries:
                if fid in exclude_ids:
                    continue
                scored.append((score, coll, fid, label))
    scored.sort(key=lambda x: (-x[0], x[3]))
    seen, out = set(), []
    for score, coll, fid, label in scored:
        # Seuil bas volontairement : 'PORTAIL HOUX 5' vs 'PORTE HOUX 3' ne
        # partagent qu'un token sur quatre (0.25) alors que c'est la bonne
        # famille. L'utilisateur tranche, mieux vaut proposer que taire.
        if fid in seen or score < 0.2:
            continue
        seen.add(fid)
        out.append({'collection': coll, '_id_feature': fid, 'label': label,
                    'score': round(score, 2)})
        if len(out) >= limit:
            break
    return out


def _candidates_for(unit, geo):
    """Rapprochements normalises possibles : [(collection, fid, label)]."""
    found = []
    if unit['kind'] == 'porte':
        for fid, label in geo['portes'].get(canonical_porte(unit['name']), []):
            found.append(('portes', fid, label))
        return found

    num = _tribune_number(unit['name'])
    if num:
        for key in _tribune_keys(num):
            for fid, label in geo['_tribune_by_num'].get(key, []):
                if (fid, label) not in [(f, l) for _, f, l in found]:
                    found.append(('tribunes', fid, label))
        if found:
            return found

    for variant in _zone_variants(unit['name']):
        for coll in ('terrains', 'hospitalites', 'tribunes'):
            for fid, label in geo[coll].get(variant, []):
                if fid not in [f for _, f, _ in found]:
                    found.append((coll, fid, label))
    return found


def resolve_features(db, units, event=None, year=None, overrides=None):
    """Rattache chaque unite a une feature geographique.

    Trois niveaux, du plus fiable au moins fiable :
      1. corrections manuelles persistees (`scan_feature_overrides`)
      2. recolte des documents `portes` deja en base (portes uniquement)
      3. rapprochement normalise contre le geojson

    `overrides` (optionnel) : mapping ad hoc fourni par l'UI pour cet import,
    au format {(kind, name): {_id_feature, feature_collection}} — prioritaire
    sur tout le reste.

    Retourne {(kind, name): {_id_feature, feature_collection, feature_source,
                             feature_label, candidates}}.
    `feature_source` vaut 'manuel', 'historique', 'geojson' ou 'aucun'.
    """
    geo = _load_geo_index(db)
    known = _harvest_known_doors(db)
    saved = _load_overrides(db, event, year)
    live = overrides or {}

    result = {}
    for unit in units:
        key = (unit['kind'], unit['name'])
        norm = normalize_label(unit['name'])
        harvest_key = canonical_porte(unit['name'])
        candidates = _candidates_for(unit, geo)
        entry = {
            '_id_feature': None,
            'feature_collection': None,
            'feature_source': 'aucun',
            'feature_label': None,
            'candidates': [{'collection': c, '_id_feature': f, 'label': l}
                           for c, f, l in candidates],
        }

        # Fusion, pas remplacement : un choix de l'UI qui ne porte QUE sur la
        # categorie ne doit pas effacer un rattachement geographique deja
        # memorise (et reciproquement).
        manual = dict(saved.get((unit['kind'], norm)) or {})
        manual.update({k: v for k, v in (live.get(key) or {}).items()
                       if v is not None})
        if manual.get('_id_feature'):
            entry.update({
                '_id_feature': manual['_id_feature'],
                'feature_collection': manual.get('feature_collection'),
                'feature_source': 'manuel',
            })
        elif unit['kind'] == 'porte' and harvest_key in known:
            entry.update({
                '_id_feature': known[harvest_key],
                'feature_collection': 'portes',
                'feature_source': 'historique',
            })
        elif len(candidates) == 1:
            coll, fid, label = candidates[0]
            entry.update({
                '_id_feature': fid,
                'feature_collection': coll,
                'feature_source': 'geojson',
                'feature_label': label,
            })
        # plusieurs candidats sans arbitrage : on laisse non resolu et on
        # expose `candidates` pour que l'utilisateur tranche dans l'UI.

        if entry['_id_feature'] and not entry['feature_label']:
            for c, f, l in candidates:
                if f == entry['_id_feature']:
                    entry['feature_label'] = l
                    break

        if not entry['_id_feature']:
            entry['suggestions'] = _suggest_for(
                unit, geo, exclude_ids={c['_id_feature'] for c in entry['candidates']})
        else:
            entry['suggestions'] = []

        # Categorie : choix explicite s'il existe, proposition sinon. Elle est
        # independante du rattachement geographique — une unite peut etre
        # localisee sans qu'on sache la classer, et l'inverse.
        chosen = (manual or {}).get('category')
        if chosen in CATEGORY_IDS:
            entry['category'] = chosen
            entry['category_source'] = 'manuel'
        else:
            entry['category'] = guess_category(
                unit['name'], unit['kind'], entry['feature_collection'])
            entry['category_source'] = 'auto'

        # Unite ecartee de l'import : elle ne figurera dans AUCUN des trois
        # documents. Jamais deduit — c'est une decision d'exploitation, et
        # ecarter une porte modifie la serie de l'enceinte generale.
        entry['ignored'] = bool(manual.get('ignored'))

        result[key] = entry
    return result


def ignored_keys(resolved):
    """Cles des unites ecartees, au format `unit['key']`."""
    return {k for k, v in (resolved or {}).items() if v.get('ignored')}


# ---------------------------------------------------------------------------
# Construction des documents historique_controle
# ---------------------------------------------------------------------------

# Plafond de securite sous la limite BSON de 16 Mo. Un doc `complet` pese
# ~91 octets par creneau et par unite : a 120 unites x 900 creneaux on frole
# les 10 Mo, d'ou le garde-fou plutot qu'une erreur Mongo opaque a l'ecriture.
MAX_DOC_BYTES = 12 * 1024 * 1024


def _iso(dt):
    """Datetime naif -> chaine ISO naive (heure locale Paris, convention maison)."""
    return dt.strftime('%Y-%m-%dT%H:%M:%S')


def _sorted_slots(unit):
    return sorted(unit['slots'].items())


def _hourly(unit):
    """Agrege les creneaux 15 min en creneaux horaires alignes."""
    buckets = defaultdict(lambda: {'e': 0, 's': 0, 'a': 0})
    for ts, counts in unit['slots'].items():
        hour = ts.replace(minute=0, second=0, microsecond=0)
        for k in ('e', 's', 'a'):
            buckets[hour][k] += counts[k]
    return sorted(buckets.items())


def build_complet_doc(parsed, resolved, event, year, race_iso=None,
                      imported_by=None, imported_at=None, uam_by_porte=None):
    """Document `type: "complet"` — granularite 15 min, source du rapport.

    `entree`/`sortie`/`autre` sont des deltas par creneau ; `present` est le
    cumul courant de (entree - sortie). `autre` en est exclu : un scan sans
    direction ne permet aucun solde.

    Pas de champ `id` dans `data_15min` : l'uuid n'a aucune signification
    inter-collection et pese 45 des 136 octets de chaque enregistrement, ce
    qui rapproche dangereusement le document de la limite BSON de 16 Mo.
    """
    uam_by_porte = uam_by_porte or {}
    ignored = ignored_keys(resolved)
    units_out = []
    for unit in parsed['units']:
        if unit['key'] in ignored:
            continue
        feat = resolved.get(unit['key'], {})
        present = 0
        series = []
        for ts, counts in _sorted_slots(unit):
            present += counts['e'] - counts['s']
            series.append({
                'date': _iso(ts),
                'entree': counts['e'],
                'sortie': counts['s'],
                'autre': counts['a'],
                'present': present,
            })
        renfort = uam_by_porte.get(unit['name']) or {}
        units_out.append({
            'uam_help': renfort.get('uam_help'),
            'pda_renfort': renfort.get('pda_renfort'),
            'name': unit['name'],
            'kind': unit['kind'],
            'zone': unit['zone'],
            'porte': unit['porte'],
            '_id_feature': feat.get('_id_feature'),
            'feature_collection': feat.get('feature_collection'),
            'feature_source': feat.get('feature_source', 'aucun'),
            # Categorie d'exploitation : pilote le regroupement et la capacite
            # retenue dans le rapport. Sans elle, le gabarit retombe sur les
            # prefixes de nom, conventions 24H AUTOS.
            'category': feat.get('category') or guess_category(
                unit['name'], unit['kind'], feat.get('feature_collection')),
            'category_source': feat.get('category_source', 'auto'),
            'devices': unit['devices'],
            'device_count': unit['device_count'],
            'pda_count': unit['pda_count'],
            'tripode_count': unit['tripode_count'],
            'other_count': unit['other_count'],
            'total_entree': unit['total_entree'],
            'total_sortie': unit['total_sortie'],
            'total_autre': unit['total_autre'],
            'data_15min': series,
        })

    return {
        'event': event,
        'year': int(year),
        'type': 'complet',
        'race': race_iso,
        'granularity': '15min',
        'source_file': parsed.get('source_file'),
        'imported_at': imported_at or datetime.now(),
        'imported_by': imported_by,
        'period_start': _iso(parsed['period_start']),
        'period_end': _iso(parsed['period_end']),
        'slot_count': parsed['slot_count'],
        'complet': units_out,
    }


def build_frequentation_doc(parsed, event, year, race_iso=None, resolved=None):
    """Document `type: "frequentation"` — serie globale horaire de l'enceinte.

    Shape strictement identique a l'existant : une seule serie sous la cle
    `data`, `entree`/`sortie` CUMULES (et non des deltas comme dans `complet`),
    `present = entree - sortie`. Les lecteurs (pcorg_summary._index_freq_by_day,
    app.py:7133) sont cables dessus, y compris sur la presence de `present`.

    `autre` est exclu : sans direction il fausserait le solde de presents.

    ⚠ Ecarter une porte change CE document, donc la reference N-1 de toutes
    les comparaisons du cockpit. C'est voulu quand l'unite n'est pas une porte
    d'enceinte (guichet, service), mais ce n'est jamais anodin.
    """
    ignored = ignored_keys(resolved)
    hourly = defaultdict(lambda: {'e': 0, 's': 0})
    excluded_autre = 0
    doors_without_direction = []
    for unit in parsed['units']:
        if unit['kind'] != 'porte' or unit['key'] in ignored:
            continue  # seules les portes d'enceinte composent la serie globale
        excluded_autre += unit['total_autre']
        # Seuil et non egalite stricte : PORTE ANNEXE totalise 19 389 scans
        # `Autre` pour 2 entrees, ce qui est un boitier sans direction avec
        # deux scans parasites, pas une porte instrumentee.
        total = unit['total_entree'] + unit['total_sortie'] + unit['total_autre']
        if total and unit['total_autre'] / float(total) > 0.9:
            doors_without_direction.append(unit['name'])
        for hour, counts in _hourly(unit):
            hourly[hour]['e'] += counts['e']
            hourly[hour]['s'] += counts['s']

    data = []
    cum_e = cum_s = 0
    for hour, counts in sorted(hourly.items()):
        cum_e += counts['e']
        cum_s += counts['s']
        data.append({
            'id': str(uuid.uuid4()),
            'date': _iso(hour),
            'entree': cum_e,
            'sortie': cum_s,
            'present': cum_e - cum_s,
        })

    # Tracabilite : la courbe de presents sous-compte mecaniquement de tous
    # les scans sans direction. Sur 24H MOTOS 2024, PORTE ANNEXE et
    # PORTE PANORAMA n'ont QUE des scans `Autre` : leur trafic (41 282 scans)
    # est totalement absent du solde. On l'ecrit dans le document plutot que
    # de laisser un lecteur futur le decouvrir a ses depens.
    return {
        'event': event,
        'year': int(year),
        'type': 'frequentation',
        'race': race_iso,
        'data': data,
        'source': 'scan_import',
        'excluded_autre': excluded_autre,
        'doors_without_direction': doors_without_direction,
        # Portes volontairement ecartees : sans cette trace, un lecteur futur
        # ne peut pas savoir pourquoi le total differe d'une autre generation.
        'ignored_doors': sorted(name for kind, name in ignored if kind == 'porte'),
    }


def build_portes_doc(parsed, resolved, event, year, race_iso=None):
    """Document `type: "portes"` — une entree par porte d'enceinte, horaire.

    `scan_count` compte TOUS les scans de l'heure, sens confondus (entree +
    sortie + autre) : c'est la semantique des documents deja en base, et sans
    le `autre` deux portes (ANNEXE, PANORAMA) tomberaient a zero.

    `doors_id` est le `_id_feature` ; il est omis quand l'unite n'a pas pu
    etre localisee, exactement comme PORTAIL HOUX 5 aujourd'hui.
    """
    ignored = ignored_keys(resolved)
    doors = []
    for unit in parsed['units']:
        if unit['kind'] != 'porte' or unit['key'] in ignored:
            continue
        scans = []
        for hour, counts in _hourly(unit):
            total = counts['e'] + counts['s'] + counts['a']
            if not total:
                continue
            scans.append({
                'id': str(uuid.uuid4()),
                'timestamp': hour,   # datetime BSON naif (heure locale Paris)
                'scan_count': total,
            })
        door = {'name': unit['name'], 'scans': scans}
        fid = (resolved.get(unit['key']) or {}).get('_id_feature')
        if fid:
            door['doors_id'] = fid
        doors.append(door)

    return {
        'event': event,
        'year': int(year),
        'type': 'portes',
        'race': race_iso,
        'doors': doors,
    }


# ---------------------------------------------------------------------------
# Ecriture : archivage puis remplacement
# ---------------------------------------------------------------------------

ARCHIVE_COLLECTION = 'historique_controle_archive'


def _doc_size(doc):
    import bson
    return len(bson.BSON.encode(doc))


def archive_and_replace(db, doc, archived_by=None, import_id=None):
    """Archive le document existant puis ecrit le nouveau.

    L'index unique de `historique_controle` porte sur (event, year, type) :
    on archive d'abord une copie integrale de l'ancien dans
    `historique_controle_archive`, puis on remplace. Si le remplacement echoue
    apres l'archivage, l'archive subsiste (trace) et le principal reste
    inchange : l'etat reste coherent.

    `year` DOIT etre un int — un str creerait un doublon sous l'index unique
    au lieu de mettre a jour.
    """
    size = _doc_size(doc)
    if size > MAX_DOC_BYTES:
        raise ScanImportError('document_trop_volumineux', 413,
                              'type=%s taille=%.1f Mo (max %.0f Mo)'
                              % (doc['type'], size / 1048576.0,
                                 MAX_DOC_BYTES / 1048576.0))

    key = {'event': doc['event'], 'year': doc['year'], 'type': doc['type']}
    existing = db['historique_controle'].find_one(key)
    archive_id = None
    if existing:
        snapshot = dict(existing)
        snapshot.pop('_id', None)
        snapshot['archived_at'] = datetime.now()
        snapshot['archived_by'] = archived_by
        snapshot['archived_reason'] = 'reimport'
        snapshot['replaced_by_import_id'] = import_id
        snapshot['original_id'] = existing['_id']
        archive_id = db[ARCHIVE_COLLECTION].insert_one(snapshot).inserted_id

    db['historique_controle'].replace_one(key, doc, upsert=True)
    return {'archived': existing is not None, 'archive_id': archive_id,
            'bytes': size}


def ensure_indexes(db):
    """Index crees paresseusement au premier import."""
    db[ARCHIVE_COLLECTION].create_index(
        [('event', 1), ('year', -1), ('type', 1), ('archived_at', -1)])
    db[OVERRIDES_COLLECTION].create_index(
        [('scan_name', 1), ('kind', 1), ('event', 1), ('year', 1)], unique=True)


def save_overrides(db, overrides, created_by=None):
    """Persiste les choix manuels pour les reutiliser aux imports suivants.

    Un choix de categorie seul est memorisable : classer une zone ne suppose
    pas de savoir ou elle se trouve.
    """
    saved = 0
    for (kind, name), val in (overrides or {}).items():
        if not val.get('save'):
            continue
        fields = {'updated_at': datetime.now(), 'created_by': created_by}
        if val.get('_id_feature'):
            fields['_id_feature'] = val['_id_feature']
            fields['feature_collection'] = val.get('feature_collection')
        if val.get('category') in CATEGORY_IDS:
            fields['category'] = val['category']
        # `ignored` est memorise dans les deux sens : reactiver une unite
        # ecartee doit survivre au prochain import, sinon elle disparaitrait
        # a nouveau sans que personne ne comprenne pourquoi.
        if val.get('ignored') is not None:
            fields['ignored'] = bool(val['ignored'])
        if len(fields) == 2:  # rien d'autre que l'horodatage
            continue
        db[OVERRIDES_COLLECTION].update_one(
            {'scan_name': name, 'kind': kind,
             'event': val.get('event'), 'year': val.get('year')},
            {'$set': fields},
            upsert=True)
        saved += 1
    return saved


# ---------------------------------------------------------------------------
# Date de course
# ---------------------------------------------------------------------------

def resolve_race(db, event, year):
    """Date de course au format ISO naif, ou None.

    Ordre de resolution, du plus fiable au moins fiable :
      1. le document `historique_controle` deja en base (portes puis
         frequentation) — souvent la seule source pour les editions passees
      2. `parametrages.data.globalHoraires.race` puis `parametrages.data.race`
         (l'annee y est stockee en STR en prod, on tente les deux formes)
    Si rien ne sort, l'appelant doit demander la date a l'utilisateur.
    """
    for hc_type in ('portes', 'frequentation'):
        for y in (int(year), str(year)):
            doc = db['historique_controle'].find_one(
                {'type': hc_type, 'event': event, 'year': y}, {'race': 1})
            if doc and doc.get('race'):
                return doc['race']

    proj = {'data.race': 1, 'data.globalHoraires.race': 1}
    for y in (str(year), int(year)):
        doc = db['parametrages'].find_one({'event': event, 'year': y}, proj)
        if not doc:
            continue
        data = doc.get('data') or {}
        for raw in ((data.get('globalHoraires') or {}).get('race'), data.get('race')):
            if raw:
                return str(raw)
    return None


# ---------------------------------------------------------------------------
# Modele xlsx telechargeable
# ---------------------------------------------------------------------------

def build_template_workbook(db=None):
    """Classeur d'exemple au format attendu, retourne un BytesIO.

    Genere a la volee (pas de fichier sur disque a maintenir). La feuille
    « Scans » reproduit les quatre largeurs de groupe reellement observees
    dans les exports Sirius, y compris le groupe a trois colonnes avec
    `Autre` en tete — c'est le cas que les anciens scripts d'import
    laissaient tomber silencieusement.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Scans'

    base = datetime(2024, 4, 15, 7, 30)
    # (offset minutes, sens presents) — les 4 largeurs de groupe observees
    groups = [
        (0,  ['Autre', 'Entrée', 'Sortie']),
        (15, ['Entrée', 'Sortie']),
        (30, ['Entrée']),
        (45, ['Sortie']),
        (60, ['Entrée', 'Sortie']),
    ]

    ws.cell(row=2, column=1, value='Zone')
    ws.cell(row=2, column=2, value='Porte')
    ws.cell(row=2, column=3, value='Device')

    col = 4
    for offset, senses in groups:
        ts = base + timedelta(minutes=offset)
        start = col
        ws.cell(row=1, column=start, value=ts).number_format = 'yyyy-mm-dd hh:mm'
        for sense in senses:
            ws.cell(row=2, column=col, value=sense)
            ws.cell(row=3, column=col, value='SPACE_CODE - Identifiant')
            col += 1
        if len(senses) > 1:
            ws.merge_cells(start_row=1, start_column=start,
                           end_row=1, end_column=col - 1)
    total_col = col
    ws.cell(row=1, column=total_col, value='Total')

    # Deux zones : l'enceinte (= les portes) et un parking, avec report vers
    # le bas (zone et porte ne sont ecrites que sur la 1re ligne du bloc).
    sample = [
        ('ENCEINTE GENERALE', 'PORTE SUD', 'PDA.101', [3, 12, 4, 9, 2, 7, 5, 6]),
        (None, None, 'TRI-SUD-01', [0, 5, 1, 4, 0, 3, 2, 2]),
        (None, 'PORTE NORD PIETONS', 'PDA.71', [0, 31, 8, 22, 6, 14, 9, 11]),
        ('PARKING OUEST', 'PORTAIL OUEST 3', 'PDA.169', [0, 8, 2, 5, 1, 4, 3, 3]),
    ]
    for i, (zone, porte, device, values) in enumerate(sample):
        row = 4 + i
        if zone:
            ws.cell(row=row, column=1, value=zone)
        if porte:
            ws.cell(row=row, column=2, value=porte)
        ws.cell(row=row, column=3, value=device)
        for j, val in enumerate(values):
            if 4 + j >= total_col:
                break
            if val:  # une cellule vide signifie « aucun scan », pas zero
                ws.cell(row=row, column=4 + j, value=val)
        ws.cell(row=row, column=total_col, value=sum(values))
    ws.cell(row=4 + len(sample), column=1, value='Total')

    for r in (1, 2, 3):
        for c in range(1, total_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=(r <= 2))
            cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    for c in range(4, total_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = 'D4'

    notice = wb.create_sheet('Notice')
    for i, line in enumerate(_TEMPLATE_NOTICE, start=1):
        cell = notice.cell(row=i, column=1, value=line)
        if line and not line.startswith(' '):
            cell.font = Font(bold=True)
    notice.column_dimensions['A'].width = 110

    if db is not None:
        _append_known_units_sheet(wb, db)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_TEMPLATE_NOTICE = [
    'Format attendu pour l\'import des scans',
    '',
    'Structure',
    '  Ligne 1 : horodatage, uniquement sur la premiere colonne de chaque groupe fusionne.',
    '  Ligne 2 : sens du scan — Entree, Sortie ou Autre.',
    '  Ligne 3 : libelle technique, ignore a l\'import.',
    '  Colonnes A/B/C : zone, porte, device. Le libelle n\'est ecrit que sur la',
    '                   premiere ligne de son bloc (report vers le bas).',
    '  Lignes 4 et suivantes : nombres entiers de scans, par pas de 15 minutes.',
    '',
    'Regles',
    '  Une cellule vide signifie « aucun scan » : ne pas la remplir avec un zero.',
    '  Un groupe peut compter 1, 2 ou 3 colonnes selon les sens presents.',
    '  La colonne Total et la ligne dont la zone vaut Total sont ignorees.',
    '  Une meme porte peut apparaitre sous deux zones differentes : la cle est',
    '  le couple (zone, porte), jamais la porte seule.',
    '',
    'Le sens « Autre »',
    '  Il ne s\'agit pas d\'une categorie de scan mais d\'un boitier non configure',
    '  en entree/sortie. Ces scans sont comptes dans le total de passages d\'une',
    '  porte, mais exclus du calcul des presents : sans direction, aucun solde',
    '  n\'est calculable.',
    '',
    'Zones et portes',
    '  La zone ENCEINTE GENERALE porte les acces du perimetre du circuit :',
    '  ses lignes deviennent les « portes ». Toutes les autres zones (tribunes,',
    '  parkings, aires d\'accueil, hospitalites) sont agregees au niveau zone.',
    '',
    'Rattachement geographique',
    '  Chaque unite est reliee a une entite cartographique via son _id_feature,',
    '  ce qui permet de comparer une meme porte d\'une edition a l\'autre meme si',
    '  son libelle change. La feuille « Unites connues » liste les libelles deja',
    '  reconnus : les reprendre a l\'identique evite l\'etape de mapping manuel.',
]


def _append_known_units_sheet(wb, db):
    """Feuille listant les libelles deja reconnus, pour aligner l'orthographe."""
    from openpyxl.styles import Font
    ws = wb.create_sheet('Unites connues')
    ws.cell(row=1, column=1, value='Source').font = Font(bold=True)
    ws.cell(row=1, column=2, value='Libelle reconnu').font = Font(bold=True)
    row = 2
    try:
        for name in sorted(_harvest_known_doors(db)):
            ws.cell(row=row, column=1, value='porte (historique)')
            ws.cell(row=row, column=2, value=name)
            row += 1
        geo = _load_geo_index(db)
        for coll in FEATURE_COLLECTIONS:
            labels = sorted({lbl for entries in geo[coll].values()
                             for _, lbl in entries})
            for lbl in labels:
                ws.cell(row=row, column=1, value=coll)
                ws.cell(row=row, column=2, value=lbl)
                row += 1
    except Exception:
        pass  # la notice reste utile meme si la base est injoignable
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 44


# ---------------------------------------------------------------------------
# Aide UAM et renforts PDA
# ---------------------------------------------------------------------------

def load_tripode_portes(db):
    """Noms normalises des portes equipees de tripodes (geojson `portes`)."""
    out = set()
    geo = db['portes'].find_one() or {}
    for f in geo.get('features', []) or []:
        p = f.get('properties') or {}
        name = p.get('Name')
        if name and (p.get('place_config') or {}).get('tripodes'):
            out.add(canonical_porte(name))
    return out


def _summarize_slots(per_hour):
    """{heure: {'pdas', 'scans'}} -> structures per_hour / per_day / total."""
    out_hour, per_day = {}, defaultdict(lambda: {'pdas': set(), 'scans': 0, 'hours': 0})
    for hour, slot in sorted(per_hour.items()):
        key = hour.strftime('%Y-%m-%dT%H:00')
        out_hour[key] = {'pda_count': len(slot['pdas']),
                         'scan_count': slot['scans']}
        day = per_day[key[:10]]
        day['pdas'].update(slot['pdas'])
        day['scans'] += slot['scans']
        day['hours'] += 1
    out_day = {d: {'pda_count': len(v['pdas']), 'scan_count': v['scans'],
                   'active_hours': v['hours']}
               for d, v in sorted(per_day.items())}
    return {
        'per_hour': out_hour,
        'per_day': out_day,
        'total_scans': sum(v['scan_count'] for v in out_hour.values()),
        'pda_count': len({p for s in per_hour.values() for p in s['pdas']}),
    }


def compute_uam_help(parsed, tripode_portes):
    """Aide UAM et renforts PDA par porte.

    Deux signaux distincts, repris de import_uam_help.py :
      - `uam_help` : un boitier de l'Unite d'Appui Mobile a scanne sur cette
        porte, donc du renfort mobile y est passe ;
      - `pda_renfort` : sur une porte a tripodes, un PDA non-UAM a servi, ce
        qui trahit un debordement des tourniquets.

    Retourne {porte: {'uam_help': {...}, 'pda_renfort': {...} | None}}.
    Sans ligne UAM dans le fichier (cas de 24H MOTOS 2024), `uam_help` est
    simplement vide partout — ce n'est pas une anomalie.
    """
    uam = set(parsed.get('uam_devices') or [])
    per_porte_uam = defaultdict(lambda: defaultdict(lambda: {'pdas': set(), 'scans': 0}))
    per_porte_renfort = defaultdict(lambda: defaultdict(lambda: {'pdas': set(), 'scans': 0}))

    for (porte, device), hours in (parsed.get('device_hours') or {}).items():
        is_uam = device in uam
        is_tripode = canonical_porte(porte) in tripode_portes
        if not is_uam and not is_tripode:
            continue
        for hour, scans in hours.items():
            if is_uam:
                slot = per_porte_uam[porte][hour]
            else:
                slot = per_porte_renfort[porte][hour]
            slot['pdas'].add(device)
            slot['scans'] += scans

    out = {}
    portes = set(per_porte_uam) | set(per_porte_renfort)
    for porte in portes:
        entry = {'uam_help': _summarize_slots(per_porte_uam.get(porte, {}))}
        if canonical_porte(porte) in tripode_portes:
            entry['pda_renfort'] = _summarize_slots(per_porte_renfort.get(porte, {}))
        out[porte] = entry
    return out
